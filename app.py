import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.styles import Border, Side, Alignment, Font
import io
import datetime
import calendar
import re
import json
import uuid
import zipfile  
from copy import copy
from streamlit_local_storage import LocalStorage
import streamlit.components.v1 as components

st.set_page_config(page_title="SRT Timesheet App", page_icon="🚂", layout="wide")

# ==========================================
# 📌 0. ข้อมูลตั้งต้นที่สำคัญมาก (Global Variables)
# ==========================================
leave_types = ["ย", "ย.", "พ", "พ.", "ป", "ป.", "ก", "ก.", "น", "น.", "ล", "ล.", "ลา", "อ", "อ."]
roles_list = ["นสน.", "ช.นสน.1", "ช.นสน.2", "เสมียน", "ประแจ", "กั้นถนนฯฉิมพลี", "กั้นถนนฯบางระมาด", "ลูกจ้าง", "อื่นๆ"]
shift_data = {
    "ว": {"text": "(06.00-18.00) น.", "hours": 4}, "ค": {"text": "(00.00-06.00)(18.00-24.00) น.", "hours": 4},
    "ว/ค": {"text": "(06.00-12.00)(18.00-24.00) น.", "hours": 4}, "ค/ว": {"text": "(00.00-06.00)(12.00-18.00) น.", "hours": 4},
    "0-12": {"text": "(00.00-12.00) น.", "hours": 4}, "00-12": {"text": "(00.00-12.00) น.", "hours": 4},
    "12-24": {"text": "(12.00-24.00) น.", "hours": 4}, "00-24": {"text": "(00.00-24.00) น.", "hours": 4},
    "(ว)": {"text": "(06.00-18.00) น.", "hours": 4}, "(ค)": {"text": "(00.00-06.00)(18.00-24.00) น.", "hours": 4},
    "(ว/ค)": {"text": "(06.00-12.00)(18.00-24.00) น.", "hours": 4}, "(ค/ว)": {"text": "(00.00-06.00)(12.00-18.00) น.", "hours": 4},
    "(0-12)": {"text": "(00.00-12.00) น.", "hours": 4}, "(00-12)": {"text": "(00.00-12.00) น.", "hours": 4},
    "(12-24)": {"text": "(12.00-24.00) น.", "hours": 4},
    "ย": {"text": "ย.", "hours": "-"}, "ย.": {"text": "ย.", "hours": "-"},
    "พ": {"text": "พ.", "hours": "-"}, "พ.": {"text": "พ.", "hours": "-"},
    "ป": {"text": "ป.", "hours": "-"}, "ป.": {"text": "ป.", "hours": "-"},
    "ก": {"text": "ก.", "hours": "-"}, "ก.": {"text": "ก.", "hours": "-"},
    "น": {"text": "น.", "hours": "-"}, "น.": {"text": "น.", "hours": "-"},
    "ล": {"text": "ล.", "hours": "-"}, "ล.": {"text": "ล.", "hours": "-"}, "ลา": {"text": "ลา", "hours": "-"},
    "อ": {"text": "อ.", "hours": "-"}, "อ.": {"text": "อ.", "hours": "-"} 
}

def get_thai_baht_text(number):
    txt_num = ["ศูนย์", "หนึ่ง", "สอง", "สาม", "สี่", "ห้า", "หก", "เจ็ด", "แปด", "เก้า"]
    txt_unit = ["", "สิบ", "ร้อย", "พัน", "หมื่น", "แสน", "ล้าน"]
    
    def read_num(num_str):
        res = ""
        length = len(num_str)
        for i, d in enumerate(num_str):
            v = int(d)
            if v == 0: continue
            p = length - i - 1
            if p == 0 and v == 1 and length > 1 and num_str[-2] != '0':
                res += "เอ็ด"
            elif p == 1 and v == 2:
                res += "ยี่สิบ"
            elif p == 1 and v == 1:
                res += "สิบ"
            else:
                res += txt_num[v] + txt_unit[p]
        return res
        
    parts = f"{round(number, 2):.2f}".split(".")
    baht, satang = parts[0], parts[1]
    
    text = "(เงิน"
    if int(baht) > 0:
        text += read_num(baht) + "บาท"
    elif int(baht) == 0 and int(satang) == 0:
        text += "ศูนย์บาท"
        
    if int(satang) > 0:
        text += read_num(satang) + "สตางค์"
    else:
        text += "ถ้วน"
    text += ")"
    return text

# ==========================================
# ✨ เวทมนตร์ CSS ปรับให้รองรับทั้งโหมดมืดและสว่าง
# ==========================================
st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Prompt:wght@300;400;500;600&display=swap');

        p, h1, h2, h3, h4, h5, h6, li, label, div.stMarkdown, input, textarea, select {
            font-family: 'Prompt', sans-serif !important;
        }

        div[data-testid="stVerticalBlock"] > div[style*="border"] {
            border-radius: 20px !important;
            box-shadow: 0 4px 15px rgba(0, 0, 0, 0.08) !important;
            padding: 20px !important;
            margin-bottom: 24px !important;
            transition: all 0.3s ease;
        }

        div.stButton > button {
            font-family: 'Prompt', sans-serif !important;
            border-radius: 12px !important;
            font-weight: 500 !important;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
        }
        div.stButton > button:hover {
            transform: translateY(-2px) !important;
            box-shadow: 0 6px 15px rgba(0, 0, 0, 0.15) !important;
        }

        div[data-testid="metric-container"] {
            border-radius: 15px !important;
            box-shadow: 0 4px 15px rgba(0, 0, 0, 0.08) !important;
            padding: 15px 20px !important;
            border-left: 5px solid #FF4B2B !important;
            transition: transform 0.2s ease, box-shadow 0.2s ease;
        }
        div[data-testid="metric-container"]:hover {
            transform: translateY(-3px);
            box-shadow: 0 8px 20px rgba(0, 0, 0, 0.12) !important;
        }

        .stTabs [data-baseweb="tab-list"] {
            gap: 10px;
            padding: 5px;
            border-radius: 12px;
        }
        .stTabs [data-baseweb="tab"] {
            border-radius: 8px;
            padding: 10px 20px;
            font-weight: 500;
        }

        .stTextInput input, .stNumberInput input, .stSelectbox div[data-baseweb="select"] {
            border-radius: 10px !important;
        }

        div[data-testid="stExpander"] {
            border-radius: 15px !important;
            box-shadow: 0 2px 8px rgba(0,0,0,0.05) !important;
        }
    </style>
""", unsafe_allow_html=True)

st.title("🚂 ระบบจัดการเวรและใบเบิกค่าตอบแทน (รฟท.)")

components.html("""
    <script>
        window.parent.addEventListener('beforeunload', function (e) {
            e.preventDefault();
            e.returnValue = 'คุณมีข้อมูลที่ยังไม่ได้บันทึก แน่ใจหรือไม่ว่าต้องการออกจากหน้านี้?';
        });
    </script>
""", height=0, width=0)

local_storage = LocalStorage()

def save_roster_to_local(df):
    if df is not None and not df.empty:
        json_str = df.to_json(orient='records')
        local_storage.setItem("srt_roster_data", json_str, key=f"ls_roster_{uuid.uuid4().hex}")

def load_roster_from_local():
    try:
        json_str = local_storage.getItem("srt_roster_data")
        if json_str:
            loaded_df = pd.read_json(io.StringIO(json_str), orient='records')
            for d in range(1, 32):
                if str(d) in loaded_df.columns:
                    loaded_df[str(d)] = loaded_df[str(d)].astype(str).replace('nan', '')
            return loaded_df
    except:
        return None
    return None

def sort_roster_by_role(df, emp_dict):
    role_to_group_map = {}
    for k, v in emp_dict.items():
        if v.get('is_regular', False):
            g = str(v.get('กลุ่ม', '')).strip()
            r = str(v.get('Role', '')).strip()
            if g and g != 'nan' and r not in role_to_group_map:
                role_to_group_map[r] = g
                
    for k, v in emp_dict.items():
        if not v.get('is_regular', False):
            r = str(v.get('Role', '')).strip()
            g = str(v.get('กลุ่ม', '')).strip()
            if g == r or not g or g == 'nan':
                if r in role_to_group_map:
                    v['กลุ่ม'] = role_to_group_map[r]

    temp_df = df.copy()
    group_order = {}
    g_idx = 0
    
    for k, v in emp_dict.items():
        if v.get('is_regular', False):
            g = str(v.get('กลุ่ม', '')).strip()
            if not g or g == 'nan':
                g = str(v.get('Role', 'อื่นๆ')).strip()
            if g not in group_order:
                group_order[g] = g_idx
                g_idx += 1
                
    def get_sort_key(row):
        name = str(row.get('ชื่อ-สกุล', '')).strip()
        role = str(row.get('Role (หน้าที่)', '')).strip()
        ukey = f"{name}_{role}"
        info = emp_dict.get(ukey, {})
        
        g = str(info.get('กลุ่ม', '')).strip()
        if not g or g == 'nan':
            g = str(info.get('Role', role)).strip()
            
        is_reg = info.get('is_regular', False)
        
        order_1 = group_order.get(g, 999) 
        order_2 = 0 if is_reg else 1      
        order_3 = row.name                
        
        return (order_1, order_2, order_3)
        
    temp_df['sort_key'] = temp_df.apply(get_sort_key, axis=1)
    temp_df = temp_df.sort_values('sort_key').reset_index(drop=True)
    temp_df['ลำดับ'] = range(1, len(temp_df) + 1)
    return temp_df.drop(columns=['sort_key'])

def parse_employee_dataframe(df_input, is_regular_worker=True):
    result_dict = {}
    for _, row in df_input.iterrows():
        name = str(row.get("รายชื่อ", "")).strip()
        if not name or name == "nan": continue
        pos = str(row.get("ตำแหน่ง", "-")).strip()
        p_clean = pos.replace(" ", "")
        
        role = pos
        if "นายสถานี" in p_clean or ("นสน" in p_clean and "ช.นสน" not in p_clean): role = "นสน."
        elif "ช.นสน.ตช" in p_clean or "ช.นสน.1" in p_clean: role = "ช.นสน.1"
        elif "ช.นสน.ตค" in p_clean or "ช.นสน.2" in p_clean: role = "ช.นสน.2"
        elif "เสมียน" in p_clean: role = "เสมียน"
        elif "ประแจ" in p_clean: role = "ประแจ"
        elif "ฉิมพลี" in p_clean: role = "กั้นถนนฯฉิมพลี"
        elif "บางระมาด" in p_clean: role = "กั้นถนนฯบางระมาด"
        elif "บริการ" in p_clean or "ลูกจ้าง" in p_clean: role = "ลูกจ้าง"
        elif "กั้นถนน" in p_clean: role = "อื่นๆ"
        else: role = "อื่นๆ"
        
        emp_id_raw = str(row.get("เลขประจำตัว", "-")) if pd.notna(row.get("เลขประจำตัว")) else "-"
        if emp_id_raw.endswith(".0"): emp_id_raw = emp_id_raw[:-2]
        
        acc1_raw = str(row.get("รหัสบัญชี", "-")) if pd.notna(row.get("รหัสบัญชี")) else "-"
        if acc1_raw.endswith(".0"): acc1_raw = acc1_raw[:-2]
        
        acc2_raw = str(row.get("รหัสบัญชี2", "-")) if pd.notna(row.get("รหัสบัญชี2")) else "-"
        if acc2_raw.endswith(".0"): acc2_raw = acc2_raw[:-2]
        
        group_raw = str(row.get("กลุ่ม", "")).strip()
        if group_raw == "nan" or not group_raw: group_raw = role
        
        unique_key = f"{name}_{role}"
        result_dict[unique_key] = {
            "ชื่อ-สกุล": name, "ตำแหน่ง": pos, 
            "เลขประจำตัว": emp_id_raw,
            "เงินเดือน": str(row.get("เงินเดือน", "-")) if pd.notna(row.get("เงินเดือน")) else "-", 
            "เรท": float(row.get("เรท 1 ชั่วโมง", 0)) if pd.notna(row.get("เรท 1 ชั่วโมง")) else 0.0,
            "เรท_4ชม": float(row.get("เรท 4 ชั่วโมง", 0)) if pd.notna(row.get("เรท 4 ชั่วโมง")) else 0.0,
            "เรท_1วัน": float(row.get("เรท 1 วัน", 0)) if pd.notna(row.get("เรท 1 วัน")) else 0.0,
            "ประเภทบัญชี": str(row.get("ประเภทบัญชี", "-")) if pd.notna(row.get("ประเภทบัญชี")) else "-", 
            "รหัสบัญชี": acc1_raw,
            "รหัสบัญชี2": acc2_raw,
            "กลุ่ม": group_raw,
            "Role": role, "is_regular": is_regular_worker 
        }
    return result_dict

# ==========================================
# 2. 🛡️ ระบบโหลดข้อมูลพนักงานแบบปลอดภัย
# ==========================================
saved_emp_json = local_storage.getItem("srt_employees_data")
saved_sub_json = local_storage.getItem("srt_sub_db_data")
saved_roster_json = local_storage.getItem("srt_roster_data")

if saved_emp_json and 'employees' not in st.session_state:
    try: st.session_state.employees = json.loads(saved_emp_json)
    except: st.session_state.employees = None

if saved_sub_json and 'sub_db' not in st.session_state:
    try: st.session_state.sub_db = json.loads(saved_sub_json)
    except: st.session_state.sub_db = None

if 'employees' not in st.session_state or not st.session_state.employees:
    if saved_emp_json:
        st.info("💡 **ตรวจพบข้อมูลตารางเวรที่คุณทำค้างไว้ในเครื่อง!**")
        if st.button("🔄 กู้คืนข้อมูลล่าสุดกลับมาทำงานต่อ", type="primary"):
            st.session_state.employees = json.loads(saved_emp_json)
            
            if saved_roster_json:
                try:
                    loaded_df = pd.read_json(io.StringIO(saved_roster_json), orient='records')
                    for d in range(1, 32):
                        if str(d) in loaded_df.columns:
                            loaded_df[str(d)] = loaded_df[str(d)].astype(str).replace('nan', '')
                    st.session_state.roster_df = loaded_df
                except: pass
                
            st.rerun()
            
    with st.container(border=True):
        st.warning("🔒 **ยังไม่มีข้อมูลพนักงานในระบบ กรุณาอัปโหลดไฟล์ 'ข้อมูล_...xlsx' เพื่อเริ่มต้น**")
        st.info("💡 **เคล็ดลับ:** หากต้องการให้ระบบจำ 'ฐานข้อมูลผู้แทน' ให้สร้าง Sheet ที่ 2 ในไฟล์ Excel และใส่รายชื่อผู้แทนเตรียมไว้ได้เลยครับ")
        uploaded_emp_file = st.file_uploader("📂 อัปโหลดไฟล์", type=["xlsx"])
        
        if uploaded_emp_file is not None:
            try:
                xls = pd.ExcelFile(uploaded_emp_file)
                df_reg = pd.read_excel(xls, sheet_name=0)
                emp_dict = parse_employee_dataframe(df_reg, is_regular_worker=True)
                sub_dict = {}
                if len(xls.sheet_names) > 1:
                    df_sub = pd.read_excel(xls, sheet_name=1)
                    sub_dict = parse_employee_dataframe(df_sub, is_regular_worker=False)

                st.session_state.employees = emp_dict
                st.session_state.sub_db = sub_dict
                
                local_storage.setItem("srt_employees_data", json.dumps(emp_dict), key=f"ls_emp_{uuid.uuid4().hex}")
                if sub_dict:
                    local_storage.setItem("srt_sub_db_data", json.dumps(sub_dict), key=f"ls_sub_{uuid.uuid4().hex}")
                    
                st.success("✅ โหลดข้อมูลพนักงาน และ ฐานข้อมูลผู้แทนสำเร็จ! กำลังเข้าสู่ระบบ...")
                st.rerun()
            except Exception as e:
                st.error(f"เกิดข้อผิดพลาดในการอ่านไฟล์: {e}")
    st.stop()

if 'roster_df' not in st.session_state:
    if saved_roster_json:
        try:
            loaded_df = pd.read_json(io.StringIO(saved_roster_json), orient='records')
            for d in range(1, 32):
                if str(d) in loaded_df.columns:
                    loaded_df[str(d)] = loaded_df[str(d)].astype(str).replace('nan', '')
            st.session_state.roster_df = loaded_df
            
            for _, row in loaded_df.iterrows():
                name = str(row['ชื่อ-สกุล']).strip()
                role = str(row['Role (หน้าที่)']).strip()
                unique_key = f"{name}_{role}"
                if unique_key not in st.session_state.employees:
                     st.session_state.employees[unique_key] = {
                            "ชื่อ-สกุล": name, "ตำแหน่ง": str(row['ตำแหน่งเบิก']).strip(), "เลขประจำตัว": "-", 
                            "เงินเดือน": "-", "เรท": 0.0, "เรท_4ชม": 0.0, "เรท_1วัน": 0.0,
                            "ประเภทบัญชี": "-", "รหัสบัญชี": "-", "รหัสบัญชี2": "-", "กลุ่ม": role, "Role": role, "is_regular": False
                     }
        except: pass
        
    if 'roster_df' not in st.session_state: 
        data = []
        for i, (key, info) in enumerate(st.session_state.employees.items()):
            row = {"ขึ้นหน้าใหม่": False, "ลำดับ": i+1, "ชื่อ-สกุล": info['ชื่อ-สกุล'], "ตำแหน่งเบิก": info['ตำแหน่ง'], "Role (หน้าที่)": info['Role']}
            for d in range(1, 32): row[str(d)] = ""
            data.append(row)
        df = pd.DataFrame(data)
        for d in range(1, 32): df[str(d)] = df[str(d)].astype(str)
        st.session_state.roster_df = sort_roster_by_role(df, st.session_state.employees)

if 'ขึ้นหน้าใหม่' not in st.session_state.roster_df.columns:
    st.session_state.roster_df.insert(0, 'ขึ้นหน้าใหม่', False)

# ==========================================
# 1. เมนูแถบด้านข้าง (Sidebar)
# ==========================================
with st.sidebar:
    st.markdown("## 🚂 เมนูหลัก")
    st.markdown("---")
    
    st.markdown("### 🔄 เริ่มต้นเดือนใหม่")
    if st.button("🗑️ ล้างข้อมูล (อัปโหลดรายชื่อใหม่)", type="primary", use_container_width=True):
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        components.html("<script>localStorage.clear(); window.parent.location.reload();</script>", height=0)
        st.stop()

    st.markdown("---")
    st.markdown("### 💾 สำรองข้อมูลตารางเวร")
    
    if st.button("🔄 ดึงตารางเวรล่าสุดจากเบราว์เซอร์", help="หากตารางเวรหายไป ให้กดปุ่มนี้เพื่อดึงกลับมา"):
        if saved_roster_json: 
            try:
                loaded_df = pd.read_json(io.StringIO(saved_roster_json), orient='records')
                for d in range(1, 32):
                    if str(d) in loaded_df.columns:
                        loaded_df[str(d)] = loaded_df[str(d)].astype(str).replace('nan', '')
                
                st.session_state.roster_df = loaded_df
                
                if "roster_table" in st.session_state:
                    del st.session_state["roster_table"]
                    
                st.rerun()
            except Exception as e:
                st.error("เกิดข้อผิดพลาดในการกู้คืน")
        else:
            st.warning("ไม่พบข้อมูลตารางเวรในเบราว์เซอร์ (อาจจะยังไม่เคยบันทึก หรือถูกล้างไปแล้ว)")

    if 'roster_df' in st.session_state and st.session_state.roster_df is not None:
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
            st.session_state.roster_df.to_excel(writer, index=False, sheet_name='Roster')
        st.download_button(
            label="📥 ดาวน์โหลดไฟล์สำรอง (.xlsx)", 
            data=buffer, 
            file_name=f"backup_roster_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx", 
            mime="application/vnd.ms-excel",
            use_container_width=True
        )

    st.markdown("<br>", unsafe_allow_html=True)
    uploaded_backup = st.file_uploader("📤 โหลดไฟล์สำรองเดิมมาทำต่อ", type=["xlsx"])
    if uploaded_backup:
        if st.button("ยืนยันโหลดตารางเวรเก่า", use_container_width=True):
            try:
                loaded_df = pd.read_excel(uploaded_backup)
                for d in range(1, 32):
                    if str(d) in loaded_df.columns:
                        loaded_df[str(d)] = loaded_df[str(d)].astype(str).replace('nan', '')
                
                for _, row in loaded_df.iterrows():
                    n = str(row.get('ชื่อ-สกุล', '')).strip()
                    r = str(row.get('Role (หน้าที่)', '')).strip()
                    if not n: continue
                    ukey = f"{n}_{r}"
                    if ukey not in st.session_state.employees:
                        st.session_state.employees[ukey] = {
                            "ชื่อ-สกุล": n, "ตำแหน่ง": str(row.get('ตำแหน่งเบิก', '')).strip(), 
                            "เลขประจำตัว": "-", "เงินเดือน": "-", "เรท": 0.0, "เรท_4ชม": 0.0, "เรท_1วัน": 0.0,
                            "ประเภทบัญชี": "-", "รหัสบัญชี": "-", "รหัสบัญชี2": "-", 
                            "กลุ่ม": r, "Role": r, "is_regular": False
                        }
                
                sorted_df = sort_roster_by_role(loaded_df, st.session_state.employees)
                st.session_state.roster_df = sorted_df
                save_roster_to_local(sorted_df)
                local_storage.setItem("srt_employees_data", json.dumps(st.session_state.employees), key=f"ls_emp_{uuid.uuid4().hex}")
                st.success("โหลดข้อมูลสำเร็จ และจัดเรียงรายชื่อให้ใหม่เรียบร้อย! 🎉")
                st.rerun()
            except Exception as e:
                st.error(f"Error: {e}")

saved_global = local_storage.getItem("srt_global_data")
default_global = {"val_13": "สิงหาคม", "year_be": 2569, "val_7": "5110/2520/2569", "val_8": "29 พ.ค. 69", "val_14": "01 ก.ค. 69", "public_holidays_dict": {}, "remarks_dict": {}}
try:
    if saved_global: 
        parsed = json.loads(saved_global)
        if "public_holidays" in parsed and isinstance(parsed["public_holidays"], list):
            parsed["public_holidays_dict"] = {str(k): "" for k in parsed["public_holidays"]}
        default_global.update(parsed)
except: pass

months_list = ["มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน", "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]
default_m = default_global.get("val_13", "สิงหาคม")
if default_m not in months_list: default_m = "สิงหาคม"
default_year = int(default_global.get("year_be", 2569))

month_idx = months_list.index(default_m) + 1
year_ce = default_year - 543
first_weekday, num_days = calendar.monthrange(year_ce, month_idx)

# ==========================================
# 📖 คู่มือการใช้งานระบบ 
# ==========================================
with st.expander("📖 คู่มือการใช้งานระบบ (คลิกเพื่ออ่านคำแนะนำ)"):
    st.markdown("""
    **1. การตั้งค่าเริ่มต้น:**
    - อัปโหลดไฟล์ `ข้อมูล_...xlsx` ในกล่องสีเหลือง เพื่อดึงฐานข้อมูลพนักงาน
    - สามารถเปลี่ยนเดือน, ปี พ.ศ., และกำหนดวันหยุดนักขัตฤกษ์ได้ในส่วน "ตั้งค่าข้อมูลส่วนกลาง"
    
    **2. การพิมพ์รหัสในตารางเวร (Master Data):**
    - **ทำงานปกติ:** `ว`, `ค`, `ว/ค`, `ค/ว`, `00-12`, `12-24`, `00-24`
    - **วันหยุดประจำสัปดาห์ที่มาทำงาน (เบิก 178):** ให้ใส่วงเล็บเปิด `(` ในวันที่เริ่มหยุด และวงเล็บปิด `)` ในวันสุดท้าย (เช่น `(ว` ... `ว)`)
    - **วันหยุดประจำสัปดาห์ที่ไม่ได้ทำงาน:** พิมพ์รหัส `ย` 
    - **วันลาพักผ่อน/ไปอบรม/ไปศาล:** พิมพ์รหัส `พ`, `อ`, `น`
    
    **3. การพิมพ์หมายเหตุอัตโนมัติ:**
    - เลื่อนไปตั้งค่าใน "📝 ตั้งค่าหมายเหตุอัตโนมัติ" (เช่น รหัส "อ") ระบบจะรวบรวมวันที่ทั้งหมด แล้วพิมพ์ลงใต้คำว่า 'หมายเหตุ' ให้แบบสวยงาม ไม่ซ้ำซ้อน
    
    **4. การจัดเรียงคนมาใหม่:**
    - หากตารางเรียงปนกัน ให้กดปุ่ม **"🗂️ จัดเรียงตารางใหม่"** ระบบจะดันคนมาใหม่ไปไว้ท้ายกลุ่มให้อัตโนมัติ
    
    **5. การส่งออกเอกสาร (Export):**
    - ไปที่ **"3. ส่งออกเอกสาร Excel สำเร็จรูป"**
    - เลือกแท็บ **"ส่งออกแบบกลุ่ม"**
    - สามารถกด **"ดูสรุปยอดเงิน"** ก่อนเพื่อตรวจสอบความถูกต้องได้
    - ระบบจะประมวลผลใบเบิก 177, 178 และรายงานปฏิบัติงานของทุกคนที่เลือก มัดรวมเป็นไฟล์ `.zip` แยกโฟลเดอร์ให้พร้อมส่งทันที!
    """)

# ==========================================
# 4. ตั้งค่าส่วนกลาง & ตารางเวร
# ==========================================
with st.container(border=True):
    st.subheader("⚙️ 1. ตั้งค่าข้อมูลส่วนกลางประจำเดือน")
    col_g1, col_g2 = st.columns(2)
    with col_g1:
        val_13 = st.selectbox("เดือน [13]", months_list, index=months_list.index(default_m))
        val_7 = st.text_input("คำสั่งแขวง [7]", default_global.get("val_7", ""))
    with col_g2:
        year_be = st.number_input("ปี พ.ศ. (สำหรับการจัดปฏิทิน)", min_value=2500, max_value=2600, value=default_year)
        val_8 = st.text_input("วันที่ลงคำสั่ง [8]", default_global.get("val_8", ""))
        val_14 = st.text_input("วันที่เซ็นเอกสารตัวย่อ [14]", default_global.get("val_14", ""))

    month_idx = months_list.index(val_13) + 1
    year_ce = year_be - 543
    first_weekday, num_days = calendar.monthrange(year_ce, month_idx)

    st.markdown("---")
    st.markdown("##### 🎉 ตั้งค่าวันหยุดนักขัตฤกษ์ประจำเดือน")
    
    old_ph_dict = default_global.get("public_holidays_dict", {})
    public_holidays_list = st.multiselect(
        "1) เลือกวันที่ตรงกับ 'วันหยุดนักขัตฤกษ์' (ใบ 178 จะถือเป็นวันหยุดพิเศษ)", 
        list(range(1, num_days + 1)),
        default=[int(k) for k in old_ph_dict.keys() if int(k) <= num_days]
    )

    ph_dict = {}
    if public_holidays_list:
        st.markdown("2) โปรดระบุชื่อวันหยุดนักขัตฤกษ์ (เพื่อแสดงในใบ 109 และ 178)")
        cols = st.columns(min(len(public_holidays_list), 4))
        for i, d in enumerate(sorted(public_holidays_list)):
            with cols[i % 4]:
                old_name = old_ph_dict.get(str(d), "")
                ph_name = st.text_input(f"ชื่อวันหยุด (วันที่ {d})", value=old_name, key=f"ph_{d}")
                ph_dict[str(d)] = ph_name

    # 📌 ส่วนตั้งค่าหมายเหตุอัตโนมัติ
    st.markdown("---")
    st.markdown("##### 📝 ตั้งค่าหมายเหตุอัตโนมัติ (จะนำไปพิมพ์รวบยอดไว้ใต้คำว่า 'หมายเหตุ')")
    st.info("💡 ตัวอย่าง: รหัส `อ`, เหตุผล `อบรมฯ`, อ้างอิงคำสั่ง `รฟ.ตร.5110/3745/2569 ลว. 04 ส.ค.69`")
    
    old_rm = default_global.get("remarks_dict", {})
    rm_list = []
    
    for k, v in old_rm.items():
        if isinstance(v, dict):
            rm_list.append({
                "รหัสในตาราง": k, 
                "เหตุผล (เช่น อบรมฯ)": v.get("reason", v.get("text", "")),
                "อ้างอิงคำสั่ง": v.get("ref", ""),
                "ลง 177": v.get("show_177", True),
                "ลง 178": v.get("show_178", True),
                "ลงรายงานฯ": v.get("show_report", False)
            })
        else:
            rm_list.append({
                "รหัสในตาราง": k, 
                "เหตุผล (เช่น อบรมฯ)": str(v),
                "อ้างอิงคำสั่ง": "",
                "ลง 177": True,
                "ลง 178": True,
                "ลงรายงานฯ": False
            })
            
    while len(rm_list) < 3:
        rm_list.append({"รหัสในตาราง": "", "เหตุผล (เช่น อบรมฯ)": "", "อ้างอิงคำสั่ง": "", "ลง 177": True, "ลง 178": True, "ลงรายงานฯ": False})
        
    df_rm = pd.DataFrame(rm_list)
    
    rm_col_config = {
        "รหัสในตาราง": st.column_config.TextColumn("รหัสในตาราง", width="small"),
        "เหตุผล (เช่น อบรมฯ)": st.column_config.TextColumn("เหตุผล (เช่น อบรมฯ)"),
        "อ้างอิงคำสั่ง": st.column_config.TextColumn("อ้างอิงคำสั่ง"),
        "ลง 177": st.column_config.CheckboxColumn("ลง 177", default=True),
        "ลง 178": st.column_config.CheckboxColumn("ลง 178", default=True),
        "ลงรายงานฯ": st.column_config.CheckboxColumn("ลงรายงานฯ", default=False)
    }
    
    edited_rm = st.data_editor(df_rm, num_rows="dynamic", use_container_width=True, hide_index=True, column_config=rm_col_config, key="remarks_editor")
    
    remarks_dict = {}
    for _, r in edited_rm.iterrows():
        code = str(r.get("รหัสในตาราง", "")).strip()
        reason = str(r.get("เหตุผล (เช่น อบรมฯ)", "")).strip()
        ref = str(r.get("อ้างอิงคำสั่ง", "")).strip()
        
        if code and code != "nan" and (reason != "nan" or ref != "nan"):
            remarks_dict[code] = {
                "reason": reason if reason != "nan" else "",
                "ref": ref if ref != "nan" else "",
                "show_177": bool(r.get("ลง 177", True)),
                "show_178": bool(r.get("ลง 178", True)),
                "show_report": bool(r.get("ลงรายงานฯ", False))
            }

    global_data = {
        "val_13": val_13, "year_be": year_be, "val_7": val_7, 
        "val_8": val_8, "val_14": val_14, "public_holidays_dict": ph_dict,
        "remarks_dict": remarks_dict 
    }
    local_storage.setItem("srt_global_data", json.dumps(global_data), key=f"ls_global_{uuid.uuid4().hex}")

# ==========================================
# 📈 Dashboard สรุปข้อมูล
# ==========================================
st.markdown("### 📊 ภาพรวมข้อมูลสถานี")
m1, m2, m3, m4 = st.columns(4)
m1.metric("👥 จำนวนพนักงาน", f"{len(st.session_state.roster_df)} คน")
m2.metric("🗓️ ประจำเดือน", f"{val_13} {year_be}")
m3.metric("📅 จำนวนวันในเดือนนี้", f"{num_days} วัน")
m4.metric("🎉 วันหยุดนักขัตฤกษ์", f"{len(ph_dict)} วัน")
st.markdown("---")

with st.container(border=True):
    col_title, col_btn = st.columns([4, 1])
    with col_title:
        st.subheader(f"🗓️ 2. จัดการตารางเวร 1-{num_days} วัน")
    with col_btn:
        if st.button("🗂️ จัดเรียงตารางใหม่", use_container_width=True, help="ดันคนมาแทนไปต่อท้ายกลุ่ม"):
            st.session_state.roster_df = sort_roster_by_role(st.session_state.roster_df, st.session_state.employees)
            save_roster_to_local(st.session_state.roster_df)
            st.rerun()

    with st.expander("👥 จัดการพนักงาน (เพิ่ม / แก้ไข / ลบ)"):
        tab_sub_db, tab_add, tab_del = st.tabs(["📥 ดึงชื่อจากฐานข้อมูลผู้แทน (Sheet 2)", "➕ พิมพ์เพิ่มรายชื่อเอง", "❌ ลบพนักงาน"])
        
        with tab_sub_db:
            if 'sub_db' in st.session_state and st.session_state.sub_db:
                st.info("💡 เลือกรายชื่อผู้แทนที่คุณเตรียมไว้ใน **Sheet ที่ 2** ของไฟล์ Excel ได้เลยครับ ระบบจะดึงเงินเดือนและรหัสบัญชีมาให้ครบถ้วน")
                sub_options = [f"{v['ชื่อ-สกุล']} ({v['Role']}) - {v['ตำแหน่ง']}" for v in st.session_state.sub_db.values()]
                selected_sub_display = st.selectbox("เลือกรายชื่อผู้แทน", sub_options)
                
                if st.button("➕ ดึงรายชื่อนี้เพิ่มลงในตารางเวร", type="primary"):
                    sel_name = selected_sub_display.split(" (")[0]
                    sel_role = selected_sub_display.split(" (")[1].split(")")[0]
                    sub_key = f"{sel_name}_{sel_role}"
                    
                    exists = any((r['ชื่อ-สกุล'] == sel_name and r['Role (หน้าที่)'] == sel_role) for _, r in st.session_state.roster_df.iterrows())
                    if exists:
                        st.warning("⚠️ พนักงานคนนี้อยู่ในตารางเวรแล้วครับ ไม่สามารถเพิ่มซ้ำได้")
                    else:
                        st.session_state.employees[sub_key] = st.session_state.sub_db[sub_key].copy()
                        
                        new_idx = len(st.session_state.roster_df) + 1
                        new_row = {"ขึ้นหน้าใหม่": False, "ลำดับ": new_idx, "ชื่อ-สกุล": sel_name, "ตำแหน่งเบิก": st.session_state.sub_db[sub_key]['ตำแหน่ง'], "Role (หน้าที่)": sel_role}
                        for d in range(1, 32): new_row[str(d)] = ""
                        
                        new_df = pd.DataFrame([new_row])
                        updated_df = pd.concat([st.session_state.roster_df, new_df], ignore_index=True)
                        st.session_state.roster_df = sort_roster_by_role(updated_df, st.session_state.employees)
                        
                        save_roster_to_local(st.session_state.roster_df)
                        local_storage.setItem("srt_employees_data", json.dumps(st.session_state.employees), key=f"ls_emp_{uuid.uuid4().hex}")
                        
                        st.success(f"✅ ดึงชื่อ {sel_name} ลงตารางเรียบร้อยแล้ว!")
                        st.rerun()
            else:
                st.warning("⚠️ ไม่พบฐานข้อมูลผู้แทน กรุณาสร้าง **Sheet ที่ 2** ในไฟล์ Excel แล้วกดปุ่มล้างข้อมูลเพื่ออัปโหลดไฟล์ใหม่ครับ")

        with tab_add:
            with st.form("add_emp_form"):
                groups_list = list(dict.fromkeys([v.get("กลุ่ม", v.get("Role", "อื่นๆ")) for v in st.session_state.employees.values()]))
                if not groups_list: groups_list = ["อื่นๆ"]
                
                c1, c2, c3 = st.columns(3)
                new_name = c1.text_input("ชื่อ-สกุล*")
                new_pos = c2.text_input("ตำแหน่งเบิก*")
                new_role = c3.selectbox("Role (หน้าที่)", roles_list)
                
                c4, c5 = st.columns(2)
                new_group = c4.selectbox("กลุ่ม (เพื่อจัดเรียงต่อท้ายคนประจำ)", groups_list)
                new_rate = c5.number_input("เรท 1 ชม. (บาท)", min_value=0.0, value=0.0)
                
                if st.form_submit_button("บันทึกข้อมูลพนักงาน", type="secondary"):
                    if new_name.strip() == "" or new_pos.strip() == "": st.error("กรุณากรอก ชื่อ และ ตำแหน่ง!")
                    else:
                        unique_key = f"{new_name}_{new_role}"
                        old_data = st.session_state.employees.get(unique_key, {})
                        old_salary = old_data.get("เงินเดือน", "-")
                        old_rate = float(old_data.get("เรท", 0.0))
                        
                        st.session_state.employees[unique_key] = {
                            "ชื่อ-สกุล": new_name, "ตำแหน่ง": new_pos, 
                            "เลขประจำตัว": old_data.get("เลขประจำตัว", "-"), 
                            "เงินเดือน": old_salary, 
                            "เรท": new_rate if new_rate > 0 else old_rate, 
                            "เรท_4ชม": old_data.get("เรท_4ชม", 0.0), 
                            "เรท_1วัน": old_data.get("เรท_1วัน", 0.0), 
                            "ประเภทบัญชี": old_data.get("ประเภทบัญชี", "-"), 
                            "รหัสบัญชี": old_data.get("รหัสบัญชี", "-"), 
                            "รหัสบัญชี2": old_data.get("รหัสบัญชี2", "-"), 
                            "กลุ่ม": new_group, 
                            "Role": new_role, "is_regular": old_data.get("is_regular", False)
                        }
                        
                        exists = any((r['ชื่อ-สกุล'] == new_name and r['Role (หน้าที่)'] == new_role) for _, r in st.session_state.roster_df.iterrows())
                        if not exists:
                            new_idx = len(st.session_state.roster_df) + 1
                            new_row = {"ขึ้นหน้าใหม่": False, "ลำดับ": new_idx, "ชื่อ-สกุล": new_name, "ตำแหน่งเบิก": new_pos, "Role (หน้าที่)": new_role}
                            for d in range(1, 32): new_row[str(d)] = ""
                            new_df = pd.DataFrame([new_row])
                            updated_df = pd.concat([st.session_state.roster_df, new_df], ignore_index=True)
                            st.session_state.roster_df = sort_roster_by_role(updated_df, st.session_state.employees)
                        save_roster_to_local(st.session_state.roster_df)
                        local_storage.setItem("srt_employees_data", json.dumps(st.session_state.employees), key=f"ls_emp_{uuid.uuid4().hex}")
                        st.success("✅ อัปเดตข้อมูลพนักงานเรียบร้อย!")
                        st.rerun()

        with tab_del:
            active_emp_options_del = [f"{r['ชื่อ-สกุล']} ({r['Role (หน้าที่)']})" for _, r in st.session_state.roster_df.iterrows()]
            emp_to_del = st.selectbox("เลือกพนักงานที่ต้องการลบออกจากตาราง", active_emp_options_del, key="del_emp_select")
            
            if st.button("❌ ยืนยันการลบพนักงาน", type="primary"):
                if emp_to_del:
                    del_name = emp_to_del.split(" (")[0]
                    del_role = emp_to_del.split(" (")[1].replace(")", "")
                    del_key = f"{del_name}_{del_role}"
                    
                    st.session_state.roster_df = st.session_state.roster_df[~((st.session_state.roster_df['ชื่อ-สกุล'] == del_name) & (st.session_state.roster_df['Role (หน้าที่)'] == del_role))]
                    
                    if del_key in st.session_state.employees:
                        del st.session_state.employees[del_key]
                        
                    st.session_state.roster_df['ลำดับ'] = range(1, len(st.session_state.roster_df) + 1)
                    
                    save_roster_to_local(st.session_state.roster_df)
                    local_storage.setItem("srt_employees_data", json.dumps(st.session_state.employees), key=f"ls_emp_{uuid.uuid4().hex}")
                    st.success(f"ลบรายชื่อ {del_name} ออกจากตารางเรียบร้อยแล้ว!")
                    st.rerun()

    # 📌 จัดเตรียมคอลัมน์ที่จะแสดงบนหน้าเว็บ (ซ่อนวันที่เกินออกให้หมด)
    column_config = {
        "ขึ้นหน้าใหม่": st.column_config.CheckboxColumn("ขึ้นหน้าใหม่", width="small"),
        "ตำแหน่งเบิก": st.column_config.TextColumn("ตำแหน่งเบิก", width="medium"), 
        "Role (หน้าที่)": st.column_config.SelectboxColumn("Role", options=roles_list, width="small")
    }
    
    display_cols = ["ขึ้นหน้าใหม่", "ตำแหน่งเบิก", "Role (หน้าที่)"]
    for d in range(1, num_days + 1):
        column_config[str(d)] = st.column_config.TextColumn(str(d), width="small")
        display_cols.append(str(d))

    with st.form("roster_data_form"):
        st.info("💡 **พิมพ์รหัสเวรได้อย่างลื่นไหลเหมือน Excel !** เมื่อคุณพิมพ์เสร็จทั้งหมดแล้ว ให้กดปุ่ม **'💾 บันทึกตารางเวร'** ด้านล่างสุด เพื่อบันทึกข้อมูลและตรวจสอบครับ")
        st.success(f"🔒 **คอลัมน์ 'ลำดับ' และ 'ชื่อ-สกุล' ถูกแช่แข็งรวมกันเป็นเสาเดียวแล้ว!** (ซ่อนวันที่เกินจากเดือน {val_13} ออกให้เพื่อความสบายตา)")
        
        display_df = st.session_state.roster_df.copy()
        
        display_df["ลำดับ_ชื่อ"] = display_df["ลำดับ"].astype(str) + ". " + display_df["ชื่อ-สกุล"]
        display_df = display_df.set_index("ลำดับ_ชื่อ")

        edited_display_df = st.data_editor(
            display_df[display_cols], 
            use_container_width=True, 
            column_config=column_config, 
            num_rows="fixed",
            key="roster_table"
        )
        
        submit_roster = st.form_submit_button("💾 บันทึกตารางเวรและตรวจสอบความถูกต้อง", type="primary", use_container_width=True)
        
        if submit_roster:
            edited_df = edited_display_df.reset_index()
            
            edited_df["ลำดับ"] = edited_df["ลำดับ_ชื่อ"].apply(lambda x: int(x.split(". ", 1)[0]) if ". " in str(x) and x.split(". ", 1)[0].isdigit() else 0)
            edited_df["ชื่อ-สกุล"] = edited_df["ลำดับ_ชื่อ"].apply(lambda x: x.split(". ", 1)[1] if ". " in str(x) else x)
            
            for d in range(num_days + 1, 32):
                edited_df[str(d)] = ""

            expected_cols = list(st.session_state.roster_df.columns)
            for c in expected_cols:
                if c not in edited_df.columns:
                    edited_df[c] = ""
            edited_df = edited_df[expected_cols]

            for d in range(1, 32):
                d_str = str(d)
                if d_str in edited_df.columns:
                    edited_df[d_str] = edited_df[d_str].fillna("").astype(str).replace(["None", "nan", "<NA>"], "")

            if not edited_df.equals(st.session_state.roster_df):
                for i in range(len(edited_df)):
                    if i < len(st.session_state.roster_df):
                        old_name = str(st.session_state.roster_df.iloc[i]['ชื่อ-สกุล']).strip()
                        old_role = str(st.session_state.roster_df.iloc[i]['Role (หน้าที่)']).strip()
                        
                        new_name = str(edited_df.iloc[i]['ชื่อ-สกุล']).strip()
                        new_role = str(edited_df.iloc[i]['Role (หน้าที่)']).strip()
                        
                        if (old_name != new_name or old_role != new_role) and old_name:
                            old_key = f"{old_name}_{old_role}"
                            new_key = f"{new_name}_{new_role}"
                            
                            if old_key in st.session_state.employees:
                                st.session_state.employees[new_key] = st.session_state.employees[old_key].copy()
                                st.session_state.employees[new_key]['ชื่อ-สกุล'] = new_name
                                st.session_state.employees[new_key]['Role'] = new_role
                                st.session_state.employees[new_key]['ตำแหน่ง'] = str(edited_df.iloc[i]['ตำแหน่งเบิก']).strip()
                                
                                del st.session_state.employees[old_key]

                st.session_state.roster_df = edited_df
                save_roster_to_local(edited_df)
                local_storage.setItem("srt_employees_data", json.dumps(st.session_state.employees), key=f"ls_emp_update_{uuid.uuid4().hex}")

            for _, row in edited_df.iterrows():
                name = str(row.get('ชื่อ-สกุล', '')).strip()
                role = str(row.get('Role (หน้าที่)', '')).strip()
                if not name: continue
                key = f"{name}_{role}"
                if key in st.session_state.employees:
                    st.session_state.employees[key]['Role'] = role
                    st.session_state.employees[key]['ชื่อ-สกุล'] = name
                    st.session_state.employees[key]['ตำแหน่ง'] = row.get('ตำแหน่งเบิก', '')
            
            st.rerun()

# ==========================================
# 🤖 🛡️ ระบบตรวจสอบความถูกต้องของตารางเวร (AI Validator)
# ==========================================
st.markdown("---")
st.markdown("#### 🤖 ระบบ AI ตรวจสอบข้อผิดพลาดของตารางเวร (ตามเงื่อนไข รฟท.)")
validator_errors = []

def get_shift_clean_for_val(row_data, day_num):
    if str(day_num) not in row_data: return ""
    val = row_data[str(day_num)]
    if pd.isna(val) or str(val).strip() in ["None", "nan", "<NA>", ""]: return ""
    s = str(val).strip().replace("(", "").replace(")", "")
    return s

for d in range(1, num_days + 1):
    day_info = {r: {'reg':[], 'sub':[]} for r in roles_list}
    
    for _, r_data in st.session_state.roster_df.iterrows():
        r_name = str(r_data.get('ชื่อ-สกุล', '')).strip()
        r_role = str(r_data.get('Role (หน้าที่)', '')).strip()
        shift_c = get_shift_clean_for_val(r_data, d)
        if not r_name or not shift_c: continue
        
        k = f"{r_name}_{r_role}"
        is_reg = st.session_state.employees.get(k, {}).get('is_regular', False)
        if r_role in day_info:
            if is_reg: day_info[r_role]['reg'].append({'name': r_name, 'shift': shift_c})
            else: day_info[r_role]['sub'].append({'name': r_name, 'shift': shift_c})

    # 📌 แก้บั๊กกรณีโหลดไฟล์ Backup แล้วระบบมองทุกคนเป็น "ผู้มาแทน"
    for r in roles_list:
        if not day_info[r]['reg'] and day_info[r]['sub']:
            day_info[r]['reg'].append(day_info[r]['sub'].pop(0))

    nsn_allowed = ["ว", "ค", "ค/ว", "ว/ค", "00-12", "12-24", "00-24", "0-12"]
    for p in day_info['นสน.']['reg'] + day_info['นสน.']['sub']:
        if p['shift'] not in leave_types and p['shift'] not in nsn_allowed:
            validator_errors.append(f"วันที่ {d}: {p['name']} (นสน.) ลงรหัส '{p['shift']}' ไม่ถูกต้อง (ต้องเป็น ว, ค, ค/ว, ว/ค, 00-12, 12-24, 00-24)")
            
    reg_nsn_shift = day_info['นสน.']['reg'][0]['shift'] if day_info['นสน.']['reg'] else ""
    for p in day_info['นสน.']['sub']:
        if p['shift'] not in leave_types:
            if reg_nsn_shift not in leave_types and reg_nsn_shift != "":
                validator_errors.append(f"วันที่ {d}: {p['name']} (มาแทน นสน.) เข้าเวรไม่ได้ เพราะนายสถานีตัวจริงไม่ได้ลา (ตัวจริงลง '{reg_nsn_shift}')")
                
    def get_active(role_str):
        return [x for x in (day_info[role_str]['reg'] + day_info[role_str]['sub']) if x['shift'] not in leave_types]
        
    active_ch1 = get_active('ช.นสน.1')
    active_ch2 = get_active('ช.นสน.2')
    
    # 📌 ปลดล็อกกฎ: ถ้า นสน. ลาหยุด หรือไม่ได้ลงเวรปกติ อนุญาตให้ ช.นสน.1 และ ช.นสน.2 เข้าเวรซ้อนกันในช่วงเช้าได้
    is_nsn_absent = (reg_nsn_shift in leave_types) or (not reg_nsn_shift)
    
    if not is_nsn_absent:
        for p1 in active_ch1:
            for p2 in active_ch2:
                s1, s2 = p1['shift'], p2['shift']
                if s1 == s2:
                    validator_errors.append(f"วันที่ {d}: {p1['name']} ({s1}) และ {p2['name']} ({s2}) เข้าเวรซ้ำกัน")
                elif s1 == "ว" and s2 != "ค":
                    validator_errors.append(f"วันที่ {d}: {p1['name']} ลง 'ว' แต่ {p2['name']} ลง '{s2}' (ต้องคู่กับ 'ค')")
                elif s1 == "ค" and s2 != "ว":
                    validator_errors.append(f"วันที่ {d}: {p1['name']} ลง 'ค' แต่ {p2['name']} ลง '{s2}' (ต้องคู่กับ 'ว')")
                elif s1 == "ว/ค" and s2 != "ค/ว":
                    validator_errors.append(f"วันที่ {d}: {p1['name']} ลง 'ว/ค' แต่ {p2['name']} ลง '{s2}' (ต้องคู่กับ 'ค/ว')")
                elif s1 == "ค/ว" and s2 != "ว/ค":
                    validator_errors.append(f"วันที่ {d}: {p1['name']} ลง 'ค/ว' แต่ {p2['name']} ลง '{s2}' (ต้องคู่กับ 'ว/ค')")

if validator_errors:
    st.error(f"พบข้อผิดพลาดในตารางเวรทั้งหมด {len(validator_errors)} จุด (โปรดตรวจสอบและแก้ไขก่อนกดบันทึก)")
    for err in validator_errors:
        st.warning(err, icon="⚠️")
else:
    st.success("✅ ตารางเวรถูกต้องตามเงื่อนไขของ รฟท. (ไม่พบการลงเวลาทับซ้อนหรือผิดคู่)")

# ==========================================
# 5. ฟังก์ชันสร้างไฟล์ Excel
# ==========================================

def format_days_to_ranges_text(days_list):
    if not days_list: return ""
    days_list = sorted(list(set(days_list)))
    ranges = []
    start = days_list[0]
    prev = days_list[0]
    for d in days_list[1:]:
        if d == prev + 1:
            prev = d
        else:
            ranges.append(f"{start}-{prev}" if start != prev else f"{start}")
            start = d
            prev = d
    ranges.append(f"{start}-{prev}" if start != prev else f"{start}")
    return ",".join(ranges)

def extract_employee_stats(roster_row):
    weekly_worked_holidays = [] 
    weekly_rest_holidays = []   
    leave_days_vacation = []
    
    is_in_period = False
    for d in range(1, 32):
        val = str(roster_row.get(str(d), "")).strip()
        if "(" in val: is_in_period = True
        clean_val = val.replace("(", "").replace(")", "")
        
        if clean_val in ['ย', 'ย.']:
            weekly_rest_holidays.append(d)
        elif is_in_period and clean_val and clean_val not in ["พ", "พ.", "ป", "ป.", "ก", "ก.", "น", "น.", "ล", "ล.", "ลา", "-"]:
            weekly_worked_holidays.append(d)
            
        if clean_val in ['พ', 'พ.']:
            leave_days_vacation.append(d)
            
        if ")" in val: is_in_period = False

    val_4 = format_days_to_ranges_text(weekly_rest_holidays) if weekly_rest_holidays else "-"    
    val_5 = format_days_to_ranges_text(weekly_worked_holidays) if weekly_worked_holidays else "-" 
    
    total_weekly_holidays = len(weekly_worked_holidays) + len(weekly_rest_holidays)
    val_17 = f"{total_weekly_holidays:02d}"
    
    val_9 = format_days_to_ranges_text(leave_days_vacation) if leave_days_vacation else "-"
    val_10 = str(leave_days_vacation[0]) if leave_days_vacation else "-"
    val_11 = str(leave_days_vacation[-1]) if leave_days_vacation else "-"
    val_12 = f"{len(leave_days_vacation):02d}" if leave_days_vacation else "00"
    
    return val_4, val_5, val_17, val_9, val_10, val_11, val_12

def generate_109(global_vars, roster_df, num_days, first_weekday):
    try: wb = openpyxl.load_workbook("109เปล่า.xlsx")
    except: return None, 0
    pages, current_page_rows = [], []
    for idx, row in roster_df.iterrows():
        if len(current_page_rows) >= 15 or (row.get("ขึ้นหน้าใหม่", False) and len(current_page_rows) > 0):
            pages.append(current_page_rows)
            current_page_rows = []
        current_page_rows.append(row)
    if current_page_rows: pages.append(current_page_rows)
    total_pages = max(len(pages), 1)
    if len(pages) == 0: pages = [[]]
    template_ws = wb.active
    template_ws.title = "หน้าที่ 1"
    worksheets = [template_ws]
    for p in range(2, total_pages + 1):
        new_ws = wb.copy_worksheet(template_ws)
        new_ws.title = f"หน้าที่ {p}"
        worksheets.append(new_ws)
    replacements_109 = {"[14]": global_vars["val_14"], "[13]": global_vars["val_13"], "[8]": global_vars["val_8"], "[7]": global_vars["val_7"]}
    days_th_abbr = ["จ.", "อ.", "พ.", "พฤ.", "ศ.", "ส.", "อา."]
    
    ph_dict_local = global_vars.get("public_holidays_dict", {})
    ph_append_str = ""
    if ph_dict_local:
        ph_texts = []
        for d_str, name in sorted(ph_dict_local.items(), key=lambda x: int(x[0])):
            ph_texts.append(f"วันที่ {d_str} {name}" if name else f"วันที่ {d_str}")
        ph_append_str = " (" + ", ".join(ph_texts) + ")"

    for page_idx, ws in enumerate(worksheets):
        page_num = page_idx + 1
        page_data = pages[page_idx]
        for r in range(1, 15):
            for c in range(1, 40):
                c_cell = ws.cell(row=r, column=c)
                val = c_cell.value
                if val and isinstance(val, str):
                    new_val = val
                    for k, v in replacements_109.items(): new_val = new_val.replace(k, str(v))
                    if "หน้า" in new_val and "/" in new_val: new_val = re.sub(r'หน้า\s*\d+\s*/\s*\d+', f'หน้า {page_num}/{total_pages}', new_val)
                    if type(c_cell).__name__ != 'MergedCell': c_cell.value = new_val
                    
        for r in range(35, 55):
            for c in range(1, 40):
                c_cell = ws.cell(row=r, column=c)
                val = c_cell.value
                if val and isinstance(val, str):
                    new_val = val
                    for k, v in replacements_109.items(): 
                        if k in new_val:
                            new_val = new_val.replace(k, str(v))
                    
                    if "วันหยุดนักขัตฤกษ์" in new_val and ph_append_str:
                        if ph_append_str not in new_val:
                            new_val = new_val.replace("วันหยุดนักขัตฤกษ์", f"วันหยุดนักขัตฤกษ์{ph_append_str}")
                            
                    if type(c_cell).__name__ != 'MergedCell': 
                        c_cell.value = new_val
        
        date_row = None
        for r in range(4, 10):
            if str(ws.cell(row=r, column=3).value).strip() == "1" and str(ws.cell(row=r, column=4).value).strip() == "2":
                date_row = r; break
        if not date_row: date_row = 7
        day_row = date_row - 1
        
        for r in range(8, 37, 2):
            ws.cell(row=r, column=1).value = ""
            ws.cell(row=r, column=2).value = ""
            ws.cell(row=r+1, column=2).value = ""
            for d in range(1, 32): ws.cell(row=r, column=2+d).value = ""
            
        for d in range(1, 32):
            col = 2 + d
            if d <= num_days:
                wd = (first_weekday + d - 1) % 7
                ws.cell(row=day_row, column=col).value = days_th_abbr[wd]
            else:
                ws.cell(row=day_row, column=col).value = ""
                ws.cell(row=date_row, column=col).value = ""
        
        current_excel_row = 8
        for row_data in page_data:
            c_name = ws.cell(row=current_excel_row, column=1)
            c_name.value = row_data['ลำดับ']
            
            c_name = ws.cell(row=current_excel_row, column=2)
            c_name.value = str(row_data['ชื่อ-สกุล']).strip()
            al = c_name.alignment
            c_name.alignment = Alignment(horizontal=al.horizontal if al else 'left', vertical=al.vertical if al else 'center', wrap_text=False, shrink_to_fit=True)
            
            c_pos = ws.cell(row=current_excel_row+1, column=2)
            c_pos.value = str(row_data['ตำแหน่งเบิก']).strip()
            al2 = c_pos.alignment
            c_pos.alignment = Alignment(horizontal=al2.horizontal if al2 else 'left', vertical=al2.vertical if al2 else 'center', wrap_text=False, shrink_to_fit=True)
            
            role_val = str(row_data.get('Role (หน้าที่)', '')).strip()
            
            for d in range(1, 32):
                col = 2 + d
                c_cell = ws.cell(row=current_excel_row, column=col)
                if d <= num_days:
                    shift = row_data.get(str(d), "")
                    c_cell.value = str(shift).strip() if pd.notna(shift) else ""
                    if c_cell.font: new_font = copy(c_cell.font)
                    else: new_font = Font()
                    
                    if shift:
                        s_clean = str(shift).strip().replace("(", "").replace(")", "")
                        if s_clean in leave_types: new_font.color = "FF0000" 
                        elif s_clean in ["00-12", "0-12", "12-24"]:
                            new_font.color = "008000" 
                            if new_font.size: new_font.size = new_font.size - 2 
                            else: new_font.size = 12
                        elif role_val == "กั้นถนนฯฉิมพลี" and s_clean in ["ว", "ค", "ว/ค", "ค/ว"]:
                            new_font.color = "0000FF" 
                        else: new_font.color = "000000" 
                    else: new_font.color = "000000"
                    c_cell.font = new_font
                else:
                    c_cell.value = ""
                    if c_cell.font:
                        nf = copy(c_cell.font)
                        nf.color = "000000"
                        c_cell.font = nf
                
            current_excel_row += 2
            
        if not ws.sheet_properties.pageSetUpPr: ws.sheet_properties.pageSetUpPr = PageSetupProperties()
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0 
        ws.page_setup.paperSize = ws.PAPERSIZE_A4; ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, total_pages

def generate_177(emp_info, roster_data, global_vars, ind_vars, num_days):
    if not emp_info: return None
    
    emp_id_str = str(emp_info.get("เลขประจำตัว", "-"))
    if emp_id_str.endswith(".0"): emp_id_str = emp_id_str[:-2]
    
    raw_salary = str(emp_info.get('เงินเดือน', '-'))
    if raw_salary != "-" and raw_salary.replace('.', '', 1).isdigit():
        salary_str = f"{float(raw_salary):,.0f}"
    else:
        salary_str = raw_salary

    try: wb = openpyxl.load_workbook("ใบ177 Update.xlsx")
    except: return None
    ws = wb.active
    
    has_1_2 = False
    for r in range(1, 15):
        for c in range(1, 40):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and "[1_2]" in val:
                has_1_2 = True
                break
        if has_1_2: break

    pos_full = str(emp_info.get("ตำแหน่ง", "-"))
    pos1, pos2 = pos_full, ""
    if has_1_2:
        if "ทำหน้าที่" in pos_full and len(pos_full) > 15:
            parts = pos_full.split("ทำหน้าที่", 1)
            pos1 = parts[0].strip()
            pos2 = "ทำหน้าที่" + parts[1].strip()
        elif len(pos_full) > 40:
            spaces = [m.start() for m in re.finditer(' ', pos_full)]
            if spaces:
                middle = len(pos_full) // 2
                closest_space = min(spaces, key=lambda x: abs(x - middle))
                pos1 = pos_full[:closest_space].strip()
                pos2 = pos_full[closest_space:].strip()
    
    remarks_dict = global_vars.get("remarks_dict", {}) 
    
    replacements = {
        "[NAME]": emp_info["ชื่อ-สกุล"], "[16]": emp_info.get("รหัสบัญชี", "-"), "[15]": emp_info["ประเภทบัญชี"],
        "[14]": global_vars["val_14"], "[13]": global_vars["val_13"], 
        "[17]": ind_vars.get("val_17", ""), "[12]": ind_vars.get("val_12", ""),
        "[11]": ind_vars.get("val_11", ""), "[10]": ind_vars.get("val_10", ""), "[9]": ind_vars.get("val_9", ""),
        "[8]": global_vars["val_8"], "[7]": global_vars["val_7"], "[6]": ind_vars.get("val_6", ""),
        "[5]": ind_vars.get("val_5", ""), "[4]": ind_vars.get("val_4", ""),
        "[3]": salary_str,  
        "[2]": emp_id_str, 
        "[1]": pos1, 
        "[1_2]": pos2 
    }
    
    for r in range(1, 60):
        for c in range(1, 40): 
            c_cell = ws.cell(row=r, column=c)
            val = c_cell.value
            if val and isinstance(val, str) and "[" in val:
                new_val = val
                for k, v in replacements.items(): new_val = new_val.replace(k, str(v))
                if type(c_cell).__name__ != 'MergedCell': 
                    c_cell.value = new_val
                    if "[1]" in val or "[NAME]" in val or "[1_2]" in val:
                        al = c_cell.alignment
                        c_cell.alignment = Alignment(
                            horizontal=al.horizontal if al else 'center',
                            vertical=al.vertical if al else 'center',
                            wrap_text=False,
                            shrink_to_fit=True
                        )
                
    # 📌 ระบบเรดาร์ 177 ค้นหาเลข 1
    start_row = 7
    for r in range(5, 15): 
        val1 = str(ws.cell(row=r, column=1).value).strip()
        if val1 == "1" or val1 == "1.0":
            start_row = r
            break
            
    offset = start_row - 7 
    
    rate_val = float(emp_info.get("เรท", 0.0))
    rate_4_val = float(emp_info.get("เรท_4ชม", 0.0))
    rate_1d_val = float(emp_info.get("เรท_1วัน", 0.0))
    rate_baht = int(rate_val)
    rate_satang = int(round((rate_val - rate_baht) * 100))
    
    sum_hours = 0 
    grand_total_money = 0.0
    code_days_177 = {}
    
    def set_cell_val_color(r, c, val, color_hex="000000"):
        cell = ws.cell(row=r, column=c)
        if type(cell).__name__ != 'MergedCell':
            cell.value = val
            if cell.font:
                nf = copy(cell.font)
                nf.color = color_hex
                cell.font = nf
            else:
                cell.font = Font(color=color_hex)
    
    for day in range(1, 32):
        row = start_row + day - 1
        
        if day > num_days:
            set_cell_val_color(row, 1, "", "000000") 
            set_cell_val_color(row, 2, "", "000000") 
            for col in range(3, 8): set_cell_val_color(row, col, "", "000000") 
            continue
            
        shift_raw = str(roster_data.get(str(day), "")).strip()
        shift_clean = shift_raw.replace("(", "").replace(")", "")
        sData = shift_data.get(shift_clean)
        
        is_holiday = shift_clean in leave_types
        font_color = "FF0000" if is_holiday else "000000"
        
        remark_data = remarks_dict.get(shift_clean)
        if isinstance(remark_data, dict) and remark_data.get("show_177", True):
            code_days_177.setdefault(shift_clean, []).append(day)
        
        if sData and sData["hours"] != "-":
            hours_val = int(sData["hours"])
            sum_hours += hours_val 
            
            if hours_val == 4 and rate_4_val > 0:
                total_money = rate_4_val
            elif hours_val == 8 and rate_1d_val > 0:
                total_money = rate_1d_val
            else:
                total_money = hours_val * rate_val
                
            grand_total_money += total_money
            
            total_baht = int(total_money)
            total_satang = int(round((total_money - total_baht) * 100))

            set_cell_val_color(row, 2, sData["text"], font_color)
            set_cell_val_color(row, 3, hours_val, font_color)
            set_cell_val_color(row, 4, rate_baht if rate_baht > 0 else 0, font_color)
            set_cell_val_color(row, 5, f"{rate_satang:02d}", font_color)
            set_cell_val_color(row, 6, total_baht if total_baht > 0 else 0, font_color)
            set_cell_val_color(row, 7, f"{total_satang:02d}", font_color)
        else:
            val = sData["text"] if sData else (shift_raw if shift_raw else "-")
            set_cell_val_color(row, 2, val, font_color)
            for col in range(3, 8): 
                set_cell_val_color(row, col, "-", font_color)
                    
    grand_total_baht = int(grand_total_money)
    grand_total_satang = int(round((grand_total_money - grand_total_baht) * 100))
    
    baht_text_str = get_thai_baht_text(grand_total_money)

    for r in range(35, 65):
        cell_v1 = str(ws.cell(row=r, column=1).value).strip()
        cell_v2 = str(ws.cell(row=r, column=2).value).strip()
        if "รวม" in cell_v1 or "รวม" in cell_v2:
            set_cell_val_color(r, 3, sum_hours, "000000")
            set_cell_val_color(r, 4, rate_baht if rate_baht > 0 else 0, "000000")
            set_cell_val_color(r, 5, f"{rate_satang:02d}", "000000")
            set_cell_val_color(r, 6, grand_total_baht if grand_total_baht > 0 else 0, "000000")
            set_cell_val_color(r, 7, f"{grand_total_satang:02d}", "000000")
            
            cell_txt = ws.cell(row=r+1, column=2)
            if type(cell_txt).__name__ != 'MergedCell':
                cell_txt.value = baht_text_str
            else:
                cell_txt.value = baht_text_str
            
            if cell_txt.font:
                nf = copy(cell_txt.font)
                nf.color = "000000"
                cell_txt.font = nf
            else:
                cell_txt.font = Font(color="000000")
            break
            
    remarks_to_print_177 = []
    month_abbr = ind_vars.get("val_6", "")
    for code, days in code_days_177.items():
        date_str = format_days_to_ranges_text(days)
        rm_info = remarks_dict[code]
        reason = rm_info.get("reason", "")
        ref = rm_info.get("ref", "")
        
        line1 = f"วันที่ {date_str} {month_abbr} {reason}".strip()
        remarks_to_print_177.append(line1)
        if ref:
            remarks_to_print_177.append(ref)
            
    remark_row_start = 21 + offset 
    remark_col = 8
    for c in [8, 9, 10, 11, 12]:
        found = False
        for r in range(10, 50):
            val = ws.cell(row=r, column=c).value
            if val and isinstance(val, str) and "หมายเหตุ" in val.replace(" ", ""):
                remark_row_start = r
                remark_col = c
                found = True
                break
        if found: break
            
    for i, text in enumerate(remarks_to_print_177):
        set_cell_val_color(remark_row_start + 1 + i, remark_col, text, "000000")

    if not ws.sheet_properties.pageSetUpPr: ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1; ws.page_setup.fitToWidth = 1
    ws.page_setup.paperSize = ws.PAPERSIZE_A4; ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_178(emp_info, roster_data, global_vars, ind_vars, num_days):
    if not emp_info: return None
    
    emp_id_str = str(emp_info.get("เลขประจำตัว", "-"))
    if emp_id_str.endswith(".0"): emp_id_str = emp_id_str[:-2]
    
    raw_salary = str(emp_info.get('เงินเดือน', '-'))
    if raw_salary != "-" and raw_salary.replace('.', '', 1).isdigit():
        salary_str = f"{float(raw_salary):,.0f}"
    else:
        salary_str = raw_salary

    # 📌 อัปเดตให้ดึงไฟล์ "178 อัพเดท.xlsx" เป็นหลัก ถ้าไม่เจอให้ดึง "178 อัพเดท2.xlsx" ครับ
    try: 
        wb = openpyxl.load_workbook("178 อัพเดท.xlsx")
    except: 
        try:
            wb = openpyxl.load_workbook("178 อัพเดท2.xlsx")
        except: return None
    ws = wb.active
    
    has_1_2 = False
    for r in range(1, 15):
        for c in range(1, 40):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and "[1_2]" in val:
                has_1_2 = True
                break
        if has_1_2: break

    pos_full = str(emp_info.get("ตำแหน่ง", "-"))
    pos1, pos2 = pos_full, ""
    if has_1_2:
        if "ทำหน้าที่" in pos_full and len(pos_full) > 15:
            parts = pos_full.split("ทำหน้าที่", 1)
            pos1 = parts[0].strip()
            pos2 = "ทำหน้าที่" + parts[1].strip()
        elif len(pos_full) > 40:
            spaces = [m.start() for m in re.finditer(' ', pos_full)]
            if spaces:
                middle = len(pos_full) // 2
                closest_space = min(spaces, key=lambda x: abs(x - middle))
                pos1 = pos_full[:closest_space].strip()
                pos2 = pos_full[closest_space:].strip()
                
    ph_dict_local = global_vars.get("public_holidays_dict", {})
    remarks_dict = global_vars.get("remarks_dict", {}) 
    
    replacements = {
        "[NAME]": emp_info["ชื่อ-สกุล"], "[16]": emp_info.get("รหัสบัญชี2", "-"), "[15]": emp_info["ประเภทบัญชี"],
        "[14]": global_vars["val_14"], "[13]": global_vars["val_13"],
        "[8]": global_vars["val_8"], "[7]": global_vars["val_7"],
        "[17]": ind_vars.get("val_17", ""), "[12]": ind_vars.get("val_12", ""), 
        "[11]": ind_vars.get("val_11", ""), "[10]": ind_vars.get("val_10", ""), "[9]": ind_vars.get("val_9", ""),
        "[6]": ind_vars.get("val_6", ""), "[5]": ind_vars.get("val_5", ""), "[4]": ind_vars.get("val_4", ""),
        "[3]": salary_str,  
        "[2]": emp_id_str, 
        "[1]": pos1,
        "[1_2]": pos2
    }
    
    for r in range(1, 60):
        for c in range(1, 40): 
            c_cell = ws.cell(row=r, column=c)
            val = c_cell.value
            if val and isinstance(val, str):
                new_val = val
                for k, v in replacements.items(): 
                    if k in new_val:
                        new_val = new_val.replace(k, str(v))
                if type(c_cell).__name__ != 'MergedCell': 
                    c_cell.value = new_val
                    if "[1]" in val or "[NAME]" in val or "[1_2]" in val:
                        al = c_cell.alignment
                        c_cell.alignment = Alignment(
                            horizontal=al.horizontal if al else 'center',
                            vertical=al.vertical if al else 'center',
                            wrap_text=False,
                            shrink_to_fit=True
                        )

    rate_val = float(emp_info.get("เรท", 0.0))
    daily_rate_db = float(emp_info.get("เรท_1วัน", 0.0))
    daily_rate = daily_rate_db if daily_rate_db > 0 else (rate_val * 8)
    
    # 📌 ระบบเรดาร์ 178 ค้นหาเลข 1
    start_row = 7
    for r in range(5, 15):
        val1 = str(ws.cell(row=r, column=1).value).strip()
        if val1 == "1" or val1 == "1.0":
            start_row = r - 1 # สำหรับ 178 บรรทัดที่ 1 จะถูกบวก 1 อีกทีในลูป
            break
            
    offset = start_row - 7
    
    if type(ws.cell(row=3 + offset, column=12)).__name__ != 'MergedCell': ws.cell(row=3 + offset, column=12).value = rate_val
    if type(ws.cell(row=5 + offset, column=12)).__name__ != 'MergedCell': ws.cell(row=5 + offset, column=12).value = daily_rate
        
    weekly_holiday_count = 0
    public_holiday_count = 0
    is_in_weekly_period = False
    
    def parse_holiday_string_to_set(day_str):
        holiday_set = set()
        if not day_str or day_str == "-": return holiday_set
        parts = str(day_str).split(",")
        for p in parts:
            p = p.strip()
            if "-" in p:
                try:
                    start, end = p.split("-")
                    for d in range(int(start), int(end) + 1):
                        holiday_set.add(d)
                except: pass
            elif p.isdigit():
                holiday_set.add(int(p))
        return holiday_set

    manual_weekly_holidays = parse_holiday_string_to_set(ind_vars.get('val_5', '0'))
    
    code_days_178 = {}
    
    for day in range(1, 32):
        row = start_row + day
        
        for col in range(1, 11):
            if type(ws.cell(row=row, column=col)).__name__ != 'MergedCell':
                ws.cell(row=row, column=col).value = None

        if day > num_days:
            continue 
            
        c_day = ws.cell(row=row, column=1)
        if type(c_day).__name__ != 'MergedCell':
            c_day.value = str(day) 
            
        shift_raw = str(roster_data.get(str(day), "")).strip()
        if "(" in shift_raw: is_in_weekly_period = True
        shift_clean = shift_raw.replace("(", "").replace(")", "")
        
        remark_data = remarks_dict.get(shift_clean)
        if isinstance(remark_data, dict) and remark_data.get("show_178", True):
            code_days_178.setdefault(shift_clean, []).append(day)
        
        if shift_clean and shift_clean not in leave_types and shift_clean != "-":
            is_public = str(day) in ph_dict_local
            is_weekly = day in manual_weekly_holidays
                
            if is_public or is_weekly:
                t1_start, t1_end, t2_start, t2_end = None, None, None, None
                if shift_clean == "ว": t1_start, t1_end = "06.00", "18.00"
                elif shift_clean == "ค": t1_start, t1_end, t2_start, t2_end = "00.00", "06.00", "18.00", "24.00"
                elif shift_clean == "ว/ค": t1_start, t1_end, t2_start, t2_end = "06.00", "12.00", "18.00", "24.00"
                elif shift_clean == "ค/ว": t1_start, t1_end, t2_start, t2_end = "00.00", "06.00", "12.00", "18.00"
                elif shift_clean in ["0-12", "00-12"]: t1_start, t1_end = "00.00", "12.00"
                elif shift_clean == "12-24": t1_start, t1_end = "12.00", "24.00"
                elif shift_clean == "00-24": t1_start, t1_end = "00.00", "24.00"
                else: t1_start, t1_end = shift_clean, ""
                
                c_pos = ws.cell(row=row, column=3)
                if type(c_pos).__name__ != 'MergedCell':
                    c_pos.value = emp_info["ตำแหน่ง"]
                    al_pos = c_pos.alignment
                    c_pos.alignment = Alignment(horizontal=al_pos.horizontal if al_pos else 'center', vertical=al_pos.vertical if al_pos else 'center', wrap_text=False, shrink_to_fit=True)

                if t1_start and type(ws.cell(row=row, column=4)).__name__ != 'MergedCell': ws.cell(row=row, column=4).value = t1_start
                if t1_end and type(ws.cell(row=row, column=5)).__name__ != 'MergedCell': ws.cell(row=row, column=5).value = t1_end
                if t2_start and type(ws.cell(row=row, column=6)).__name__ != 'MergedCell': ws.cell(row=row, column=6).value = t2_start
                if t2_end and type(ws.cell(row=row, column=7)).__name__ != 'MergedCell': ws.cell(row=row, column=7).value = t2_end
                if type(ws.cell(row=row, column=8)).__name__ != 'MergedCell': ws.cell(row=row, column=8).value = 1 
                
                if type(ws.cell(row=row, column=10)).__name__ != 'MergedCell': ws.cell(row=row, column=10).value = daily_rate 
                
                if is_public:
                    h_name = ph_dict_local.get(str(day), "วันหยุดนักขัตฤกษ์")
                    if not h_name.strip(): h_name = "วันหยุดนักขัตฤกษ์"
                    c_name2 = ws.cell(row=row, column=2)
                    if type(c_name2).__name__ != 'MergedCell': c_name2.value = f"({h_name})"
                    public_holiday_count += 1
                elif is_weekly:
                    c_name2 = ws.cell(row=row, column=2)
                    if type(c_name2).__name__ != 'MergedCell': c_name2.value = "(วันหยุดประจำสัปดาห์)"
                    weekly_holiday_count += 1
                    
        if ")" in shift_raw: is_in_weekly_period = False
                    
    total_days_final = weekly_holiday_count + public_holiday_count
    
    # 📌 การหาบรรทัดยอดรวม 178 อัตโนมัติ ป้องกันตารางเลื่อน
    type_acc_row = 39 + offset
    total_row = 42 + offset
    
    for r in range(start_row + 25, 60):
        val1 = str(ws.cell(row=r, column=1).value).replace(" ", "")
        if "ประเภทบัญชี" in val1:
            type_acc_row = r
            break
            
    for r in range(type_acc_row, 65):
        val1 = str(ws.cell(row=r, column=1).value).replace(" ", "")
        if "รวมเงิน" in val1:
            total_row = r
            break
            
    if type(ws.cell(row=type_acc_row, column=8)).__name__ != 'MergedCell': 
        ws.cell(row=type_acc_row, column=8).value = None
    if type(ws.cell(row=type_acc_row+1, column=8)).__name__ != 'MergedCell': 
        ws.cell(row=type_acc_row+1, column=8).value = None
    if type(ws.cell(row=type_acc_row+2, column=8)).__name__ != 'MergedCell': 
        ws.cell(row=type_acc_row+2, column=8).value = None
    
    ws.cell(row=type_acc_row, column=8).value = total_days_final
    
    if type(ws.cell(row=total_row, column=8)).__name__ != 'MergedCell':
        ws.cell(row=total_row, column=8).value = f"=SUM(H{type_acc_row}:H{total_row-1})"
        
    remarks_to_print_178 = []
    month_abbr = ind_vars.get("val_6", "")
    for code, days in code_days_178.items():
        date_str = format_days_to_ranges_text(days)
        rm_info = remarks_dict[code]
        reason = rm_info.get("reason", "")
        ref = rm_info.get("ref", "")
        
        line1 = f"วันที่ {date_str} {month_abbr} {reason}".strip()
        remarks_to_print_178.append(line1)
        if ref:
            remarks_to_print_178.append(ref)
            
    remark_row_start = 22 + offset
    remark_col = 11
    for c in [8, 9, 10, 11, 12]:
        found = False
        for r in range(10, 50):
            val = ws.cell(row=r, column=c).value
            if val and isinstance(val, str) and "หมายเหตุ" in val.replace(" ", ""):
                remark_row_start = r
                remark_col = c
                found = True
                break
        if found: break
            
    for i, text in enumerate(remarks_to_print_178):
        cell = ws.cell(row=remark_row_start + 1 + i, column=remark_col)
        if type(cell).__name__ != 'MergedCell':
            cell.value = text
            if cell.font:
                nf = copy(cell.font)
                nf.color = "000000"
                cell.font = nf
            else:
                cell.font = Font(color="000000")
    
    if not ws.sheet_properties.pageSetUpPr: ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1; ws.page_setup.fitToWidth = 1
    ws.page_setup.paperSize = ws.PAPERSIZE_A4; ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_report_work(emp_info, roster_data, global_vars, ind_vars, num_days):
    if not emp_info: return None
    
    emp_id_str = str(emp_info.get("เลขประจำตัว", "-"))
    if emp_id_str.endswith(".0"): emp_id_str = emp_id_str[:-2]
    
    raw_salary = str(emp_info.get('เงินเดือน', '-'))
    if raw_salary != "-" and raw_salary.replace('.', '', 1).isdigit():
        salary_str = f"{float(raw_salary):,.0f}"
    else:
        salary_str = raw_salary
        
    try: wb = openpyxl.load_workbook("รายงานปฏิบัติงาน.xlsx")
    except: return None
    ws = wb.active
    
    has_1_2 = False
    for r in range(1, 15):
        for c in range(1, 40):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and "[1_2]" in val:
                has_1_2 = True
                break
        if has_1_2: break

    pos_full = str(emp_info.get("ตำแหน่ง", "-"))
    pos1, pos2 = pos_full, ""
    if has_1_2:
        if "ทำหน้าที่" in pos_full and len(pos_full) > 15:
            parts = pos_full.split("ทำหน้าที่", 1)
            pos1 = parts[0].strip()
            pos2 = "ทำหน้าที่" + parts[1].strip()
        elif len(pos_full) > 40:
            spaces = [m.start() for m in re.finditer(' ', pos_full)]
            if spaces:
                middle = len(pos_full) // 2
                closest_space = min(spaces, key=lambda x: abs(x - middle))
                pos1 = pos_full[:closest_space].strip()
                pos2 = pos_full[closest_space:].strip()
    
    remarks_dict = global_vars.get("remarks_dict", {}) 

    replacements = {
        "[NAME]": emp_info["ชื่อ-สกุล"],
        "[2]": emp_id_str,
        "[3]": salary_str,
        "[14]": global_vars["val_14"],
        "[13]": global_vars["val_13"],
        "[8]": global_vars["val_8"], 
        "[7]": global_vars["val_7"],
        "[17]": ind_vars.get("val_17", ""), 
        "[1]": pos1,
        "[1_2]": pos2
    }

    for r in range(1, 60):
        for c in range(1, 40):
            c_cell = ws.cell(row=r, column=c)
            val = c_cell.value
            if val and isinstance(val, str):
                new_val = val
                for k, v in replacements.items(): new_val = new_val.replace(k, str(v))
                if type(c_cell).__name__ != 'MergedCell': 
                    c_cell.value = new_val
                    if "[1]" in val or "[NAME]" in val or "[1_2]" in val:
                        al = c_cell.alignment
                        c_cell.alignment = Alignment(
                            horizontal=al.horizontal if al else 'center',
                            vertical=al.vertical if al else 'center',
                            wrap_text=False,
                            shrink_to_fit=True
                        )

    # 📌 ระบบเรดาร์รายงานปฏิบัติงาน ค้นหาเลข 1
    start_row = 8
    for r in range(5, 15):
        val1 = str(ws.cell(row=r, column=1).value).strip()
        if val1 == "1" or val1 == "1.0":
            start_row = r
            break
            
    offset = start_row - 8

    if type(ws.cell(row=2 + min(offset, 1), column=7)).__name__ != 'MergedCell':
        ws.cell(row=2 + min(offset, 1), column=7).value = global_vars["val_13"]
    if type(ws.cell(row=44 + offset, column=2)).__name__ != 'MergedCell':
        ws.cell(row=44 + offset, column=2).value = emp_info["ชื่อ-สกุล"]

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    ranges_to_unmerge = []
    for merged_range in list(ws.merged_cells.ranges):
        if start_row <= merged_range.min_row <= start_row + 31 and start_row <= merged_range.max_row <= start_row + 31:
            if merged_range.min_col >= 2 and merged_range.max_col <= 7:
                ranges_to_unmerge.append(merged_range.coord)
    
    for r_coord in ranges_to_unmerge:
        ws.unmerge_cells(r_coord)
        
    def apply_style(r, c, val, color_hex="000000"):
        cell = ws.cell(row=r, column=c)
        if type(cell).__name__ != 'MergedCell':
            cell.value = val
            cell.border = thin_border
            if cell.font:
                nf = copy(cell.font)
                nf.color = color_hex
                cell.font = nf
            else:
                cell.font = Font(color=color_hex)
            return cell
        return None

    for day in range(1, 32):
        row = start_row + day - 1
        
        if day > num_days:
            apply_style(row, 1, "", "000000") 
            for c in range(2, 11): 
                cell = ws.cell(row=row, column=c)
                if type(cell).__name__ != 'MergedCell':
                    cell.value = None
                    cell.border = thin_border
            continue
            
        shift_raw = str(roster_data.get(str(day), "")).strip()
        start_time, end_time = "-", "-"
        is_holiday = False
        
        if shift_raw:
            s_clean = shift_raw.replace("(", "").replace(")", "")
            if s_clean == "ว": start_time, end_time = "06.00", "18.00"
            elif s_clean == "ค": start_time, end_time = "00.00-06.00", "18.00-24.00"
            elif s_clean == "ว/ค": start_time, end_time = "06.00-12.00", "18.00-24.00"
            elif s_clean == "ค/ว": start_time, end_time = "00.00-06.00", "12.00-18.00"
            elif s_clean in ["0-12", "00-12"]: start_time, end_time = "00.00", "12.00"
            elif s_clean == "12-24": start_time, end_time = "12.00", "24.00"
            elif s_clean == "00-24": start_time, end_time = "00.00", "24.00"
            elif s_clean in leave_types: 
                start_time = s_clean
                end_time = ""
                is_holiday = True 
            else: start_time, end_time = shift_raw, "" 
            
        font_color = "FF0000" if is_holiday else "000000"

        for c in range(2, 8): 
            cell = ws.cell(row=row, column=c)
            cell.value = None
            cell.border = thin_border

        if is_holiday:
            cell = apply_style(row, 2, start_time, font_color)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
            if cell: cell.alignment = center_align
        else:
            cell1 = apply_style(row, 2, start_time, font_color)
            cell2 = apply_style(row, 5, end_time, font_color)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
            if cell1: cell1.alignment = center_align
            if cell2: cell2.alignment = center_align

        apply_style(row, 8, emp_info["ตำแหน่ง"] if start_time not in ["-", ""] and not is_holiday else "", "000000")
        
        shift_clean = shift_raw.replace("(", "").replace(")", "")
        remark_data = remarks_dict.get(shift_clean)
        remark_text = ""
        if isinstance(remark_data, dict):
            if remark_data.get("show_report", False):
                reason = remark_data.get("reason", "")
                ref = remark_data.get("ref", "")
                remark_text = f"{reason} {ref}".strip()
                
        if remark_text:
            apply_style(row, 10, remark_text, "000000")
            
    if not ws.sheet_properties.pageSetUpPr: ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1; ws.page_setup.fitToWidth = 1
    ws.page_setup.paperSize = ws.PAPERSIZE_A4; ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ==========================================
# 6. เมนูส่งออก (Export)
# ==========================================
with st.container(border=True):
    st.subheader("🖨️ 3. ส่งออกเอกสาร Excel สำเร็จรูป")
    tab1, tab2, tab3 = st.tabs(["📄 ส่งออก 109", "👤 ส่งออกรายบุคคล", "📦 ส่งออกแบบกลุ่ม (หลายคนพร้อมกัน)"])

    with tab1:
        st.markdown("**ตารางเวร 109 ของทั้งสถานี** (จัดหน้ากระดาษอัตโนมัติ)")
        if st.button("คลิกเพื่อสร้างไฟล์ 109", type="primary"):
            excel_109, total_pages = generate_109(global_data, st.session_state.roster_df, num_days, first_weekday)
            if excel_109:
                st.success(f"สร้างไฟล์ 109 เสร็จสิ้น! (รวม {total_pages} หน้า)")
                st.download_button("📥 ดาวน์โหลดไฟล์ 109 (.xlsx)", data=excel_109, file_name=f"109_{global_data['val_13']}.xlsx")

    with tab2:
        st.markdown("**ข้อมูลเฉพาะบุคคลสำหรับใบเบิก**")
        st.info("💡 **ตารางเวรคือ Master Data:** หากคุณต้องการเปลี่ยนช่วงวันหยุด หรือวันลาพักผ่อน กรุณาแก้ไขเครื่องหมาย `(` `)` `ย` หรือ `พ` ในตารางเวรด้านบน ระบบจะอัปเดตกล่องข้อความด้านล่าง และซิงค์ทุกเอกสารให้อัตโนมัติ 100% ครับ")
        
        active_emp_options = [f"{r['ชื่อ-สกุล']} ({r['Role (หน้าที่)']})" for _, r in st.session_state.roster_df.iterrows()]
        selected_key_177_display = st.selectbox("เลือกพนักงานที่ต้องการสร้างเอกสาร", active_emp_options, key="single_select")
        
        if selected_key_177_display:
            sel_name = selected_key_177_display.split(" (")[0]
            sel_role = selected_key_177_display.split(" (")[1].replace(")", "")
            selected_key_177 = f"{sel_name}_{sel_role}"
            
            saved_ind = local_storage.getItem(f"srt_ind_{selected_key_177}")
            default_ind = {"val_6": "-"}
            try:
                if saved_ind: default_ind.update(json.loads(saved_ind))
            except: pass
            
            emp_row = st.session_state.roster_df[(st.session_state.roster_df['ชื่อ-สกุล'] == sel_name) & (st.session_state.roster_df['Role (หน้าที่)'] == sel_role)].iloc[0]
            roster_dict = {str(d): str(emp_row[str(d)]) if pd.notna(emp_row[str(d)]) else "" for d in range(1, 32)}
            
            val_4_auto, val_5_auto, val_17_auto, val_9_auto, val_10_auto, val_11_auto, val_12_auto = extract_employee_stats(roster_dict)

            c1, c2, c3, c4 = st.columns(4)
            with c1:
                st.text_input("วันหยุดที่ไม่ได้มาทำงาน (ย) [4]", value=val_4_auto, disabled=True)
                st.text_input("หยุดพักผ่อน [9]", value=val_9_auto, disabled=True)
            with c2:
                st.text_input("วันหยุดที่มาทำงาน (ในวงเล็บ) [5]", value=val_5_auto, disabled=True)
                st.text_input("พักผ่อนตั้งแต่ [10]", value=val_10_auto, disabled=True)
            with c3:
                st.text_input("รวมวันหยุดประจำสัปดาห์ [17]", value=val_17_auto, disabled=True)
                st.text_input("พักผ่อนถึง [11]", value=val_11_auto, disabled=True)
            with c4:
                val_6_manual = st.text_input("เดือนตัวย่อ [6]", value=default_ind.get('val_6', 'ส.ค.69'))
                st.text_input("รวมพักผ่อน [12]", value=val_12_auto, disabled=True)

            export_ind = {
                "val_4": val_4_auto, "val_5": val_5_auto, "val_17": val_17_auto,
                "val_9": val_9_auto, "val_10": val_10_auto, "val_11": val_11_auto, "val_12": val_12_auto,
                "val_6": val_6_manual
            }

            local_storage.setItem(f"srt_ind_{selected_key_177}", json.dumps({"val_6": val_6_manual}), key=f"ls_ind_{uuid.uuid4().hex}")

            st.markdown("<br>", unsafe_allow_html=True)
            col_btn1, col_btn2, col_btn3 = st.columns(3)
            
            with col_btn1:
                if st.button(f"🧾 ออกใบเบิก 177 (ทำล่วงเวลา)", use_container_width=True):
                    excel_177 = generate_177(st.session_state.employees.get(selected_key_177), roster_dict, global_data, export_ind, num_days)
                    if excel_177:
                        st.success(f"สร้างใบเบิก 177 เสร็จสิ้น!")
                        st.download_button("📥 ดาวน์โหลดไฟล์ 177", data=excel_177, file_name=f"177_{sel_name}.xlsx", use_container_width=True)
                    else:
                        st.error("ไม่พบไฟล์ 'ใบ177 Update.xlsx' ในระบบ (กรุณาอัปโหลดก่อนครับ)")
                        
            with col_btn2:
                if st.button(f"🎉 ออกใบเบิก 178 (วันหยุด)", use_container_width=True):
                    excel_178 = generate_178(st.session_state.employees.get(selected_key_177), roster_dict, global_data, export_ind, num_days)
                    if excel_178:
                        st.success(f"สร้างใบเบิก 178 เสร็จสิ้น!")
                        st.download_button("📥 ดาวน์โหลดไฟล์ 178", data=excel_178, file_name=f"178_{sel_name}.xlsx", use_container_width=True)
                    else:
                        st.error("ไม่พบไฟล์ '178 อัพเดท.xlsx' หรือ '178 อัพเดท2.xlsx' ในระบบ (กรุณาอัปโหลดก่อนครับ)")

            with col_btn3:
                if st.button(f"🕒 ออกรายงานปฏิบัติงาน", use_container_width=True):
                    excel_work = generate_report_work(st.session_state.employees.get(selected_key_177), roster_dict, global_data, export_ind, num_days)
                    if excel_work:
                        st.success(f"สร้างรายงานปฏิบัติงาน เสร็จสิ้น!")
                        st.download_button("📥 ดาวน์โหลดรายงานฯ", data=excel_work, file_name=f"รายงานปฏิบัติงาน_{sel_name}.xlsx", use_container_width=True)

            st.markdown("---")
            st.markdown("##### 📦 ดาวน์โหลดรวมทุกไฟล์ในคลิกเดียว (ZIP)")
            if st.button(f"แพ็กรวมเอกสารของ {sel_name} (177, 178, รายงาน)", type="primary", use_container_width=True):
                with st.spinner("กำลังแพ็กไฟล์..."):
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                        excel_177 = generate_177(st.session_state.employees.get(selected_key_177), roster_dict, global_data, export_ind, num_days)
                        if excel_177:
                            zip_file.writestr(f"177_{sel_name}.xlsx", excel_177.getvalue())
                            
                        excel_178 = generate_178(st.session_state.employees.get(selected_key_177), roster_dict, global_data, export_ind, num_days)
                        if excel_178:
                            zip_file.writestr(f"178_{sel_name}.xlsx", excel_178.getvalue())
                            
                        excel_work = generate_report_work(st.session_state.employees.get(selected_key_177), roster_dict, global_data, export_ind, num_days)
                        if excel_work:
                            zip_file.writestr(f"รายงานปฏิบัติงาน_{sel_name}.xlsx", excel_work.getvalue())
                            
                    st.session_state[f"zip_export_{selected_key_177}"] = zip_buffer.getvalue()

            if f"zip_export_{selected_key_177}" in st.session_state:
                st.download_button(
                    "📥 คลิกดาวน์โหลดไฟล์ ZIP ทั้งหมด",
                    data=st.session_state[f"zip_export_{selected_key_177}"],
                    file_name=f"เอกสารเบิกเงิน_{sel_name}.zip",
                    mime="application/zip",
                    use_container_width=True
                )

    with tab3:
        st.markdown("**ดาวน์โหลดเอกสารของพนักงานหลายคนพร้อมกันในคลิกเดียว ระบบจะแยกไฟล์ใส่โฟลเดอร์ให้เป็นระเบียบ**")
        
        saved_ind_batch = local_storage.getItem(f"srt_ind_batch")
        default_val_6 = "ส.ค.69"
        try:
            if saved_ind_batch: default_val_6 = json.loads(saved_ind_batch).get("val_6", "ส.ค.69")
        except: pass
        
        all_emp_options = [f"{r['ชื่อ-สกุล']} ({r['Role (หน้าที่)']})" for _, r in st.session_state.roster_df.iterrows()]
        
        select_all = st.checkbox("เลือกพนักงานทั้งหมด", value=True)
        
        with st.form("batch_export_form"):
            batch_val_6 = st.text_input("เดือนตัวย่อ [6] (ใช้กับทุกคน)", value=default_val_6, key="batch_v6")
            
            if select_all:
                selected_batch_emps = st.multiselect("เลือกพนักงานที่ต้องการส่งออก", all_emp_options, default=all_emp_options)
            else:
                selected_batch_emps = st.multiselect("เลือกพนักงานที่ต้องการส่งออก", all_emp_options)
                
            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                submit_preview = st.form_submit_button("📊 คำนวณสรุปยอดทั้งหมด", use_container_width=True)
            with col_btn2:
                submit_export = st.form_submit_button("📦 สร้างไฟล์ ZIP ส่งออกแบบกลุ่ม", type="primary", use_container_width=True)

        if submit_preview or submit_export:
            local_storage.setItem(f"srt_ind_batch", json.dumps({"val_6": batch_val_6}), key=f"ls_batch_{uuid.uuid4().hex}")

        if submit_preview:
            summary_list = []
            
            def parse_holiday_string_to_set_local(day_str):
                holiday_set = set()
                if not day_str or day_str == "-": return holiday_set
                parts = str(day_str).split(",")
                for p in parts:
                    p = p.strip()
                    if "-" in p:
                        try:
                            start, end = p.split("-")
                            for d in range(int(start), int(end) + 1):
                                holiday_set.add(d)
                        except: pass
                    elif p.isdigit():
                        holiday_set.add(int(p))
                return holiday_set

            ph_dict_local = global_data.get("public_holidays_dict", {})

            for emp_display in selected_batch_emps:
                sel_name = emp_display.split(" (")[0]
                sel_role = emp_display.split(" (")[1].replace(")", "")
                unique_key = f"{sel_name}_{sel_role}"
                emp_info = st.session_state.employees.get(unique_key)
                
                if not emp_info: continue
                
                try:
                    emp_row = st.session_state.roster_df[(st.session_state.roster_df['ชื่อ-สกุล'] == sel_name) & (st.session_state.roster_df['Role (หน้าที่)'] == sel_role)].iloc[0]
                except:
                    continue
                
                rate_val = float(emp_info.get("เรท", 0.0))
                rate_4_val = float(emp_info.get("เรท_4ชม", 0.0))
                rate_1d_val = float(emp_info.get("เรท_1วัน", 0.0))
                daily_rate = rate_1d_val if rate_1d_val > 0 else (rate_val * 8)
                
                sum_hours_177 = 0
                total_money_177 = 0.0
                for d in range(1, num_days + 1):
                    shift_raw = str(emp_row.get(str(d), "")).strip()
                    shift_clean = shift_raw.replace("(", "").replace(")", "")
                    sData = shift_data.get(shift_clean)
                    if sData and sData.get("hours", "-") != "-":
                        h = int(sData["hours"])
                        sum_hours_177 += h
                        if h == 4 and rate_4_val > 0:
                            total_money_177 += rate_4_val
                        elif h == 8 and rate_1d_val > 0:
                            total_money_177 += rate_1d_val
                        else:
                            total_money_177 += h * rate_val
                
                roster_dict = {str(d): str(emp_row[str(d)]) if pd.notna(emp_row[str(d)]) else "" for d in range(1, 32)}
                val_4, val_5, val_17, _, _, _, _ = extract_employee_stats(roster_dict)
                
                manual_weekly_holidays = parse_holiday_string_to_set_local(val_5)
                days_178 = 0
                
                for d in range(1, num_days + 1):
                    shift_raw = str(emp_row.get(str(d), "")).strip()
                    shift_clean = shift_raw.replace("(", "").replace(")", "")
                    if shift_clean and shift_clean not in leave_types and shift_clean != "-":
                        is_public = str(d) in ph_dict_local
                        is_weekly = d in manual_weekly_holidays
                        if is_public or is_weekly:
                            days_178 += 1
                            
                total_money_178 = days_178 * daily_rate
                grand_total = total_money_177 + total_money_178
                
                summary_list.append({
                    "ชื่อ-สกุล": emp_info["ชื่อ-สกุล"],
                    "ตำแหน่ง": emp_info["Role"],
                    "รวมชั่วโมง (177)": sum_hours_177,
                    "รวมเงิน 177": f"{total_money_177:,.2f}",
                    "จำนวนวัน (178)": days_178,
                    "รวมเงิน 178": f"{total_money_178:,.2f}",
                    "💰 รวมรายได้ทั้งหมด": f"{grand_total:,.2f}"
                })
                
            if summary_list:
                st.markdown("#### 📊 สรุปยอดเงินและชั่วโมง")
                st.dataframe(pd.DataFrame(summary_list), use_container_width=True)
            else:
                st.warning("ไม่มีข้อมูลให้สรุป หรือยังไม่ได้เลือกพนักงานครับ")
        
        if submit_export:
            if not selected_batch_emps:
                st.warning("กรุณาเลือกพนักงานอย่างน้อย 1 คนครับ")
            else:
                with st.spinner(f"กำลังประมวลผลเอกสาร {len(selected_batch_emps)} รายการ..."):
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                        for emp_display in selected_batch_emps:
                            sel_name = emp_display.split(" (")[0]
                            sel_role = emp_display.split(" (")[1].replace(")", "")
                            unique_key = f"{sel_name}_{sel_role}"
                            
                            emp_row = st.session_state.roster_df[(st.session_state.roster_df['ชื่อ-สกุล'] == sel_name) & (st.session_state.roster_df['Role (หน้าที่)'] == sel_role)].iloc[0]
                            roster_dict = {str(d): str(emp_row[str(d)]) if pd.notna(emp_row[str(d)]) else "" for d in range(1, 32)}
                            
                            val_4, val_5, val_17, val_9, val_10, val_11, val_12 = extract_employee_stats(roster_dict)
                            
                            export_ind = {
                                "val_4": val_4, "val_5": val_5, "val_17": val_17,
                                "val_9": val_9, "val_10": val_10, "val_11": val_11, "val_12": val_12,
                                "val_6": batch_val_6
                            }
                            
                            folder_name = f"{sel_name} {batch_val_6}"
                            
                            excel_177 = generate_177(st.session_state.employees.get(unique_key), roster_dict, global_data, export_ind, num_days)
                            if excel_177: zip_file.writestr(f"{folder_name}/177 {batch_val_6}.xlsx", excel_177.getvalue())
                            
                            excel_178 = generate_178(st.session_state.employees.get(unique_key), roster_dict, global_data, export_ind, num_days)
                            if excel_178: zip_file.writestr(f"{folder_name}/178 {batch_val_6}.xlsx", excel_178.getvalue())
                            
                            excel_work = generate_report_work(st.session_state.employees.get(unique_key), roster_dict, global_data, export_ind, num_days)
                            if excel_work: zip_file.writestr(f"{folder_name}/รายงานปฏิบัติงาน {batch_val_6}.xlsx", excel_work.getvalue())
                            
                    st.session_state["batch_zip_export"] = zip_buffer.getvalue()
                    
        if "batch_zip_export" in st.session_state:
            st.success("✅ แพ็กไฟล์สำเร็จแล้ว! กดดาวน์โหลดด้านล่างได้เลยครับ")
            st.download_button(
                "📥 ดาวน์โหลดไฟล์ ZIP (รวมทุกเอกสาร แยกแฟ้มรายบุคคลให้แล้ว)",
                data=st.session_state["batch_zip_export"],
                file_name=f"เอกสารเบิกเงิน_รฟท_{global_data['val_13']}.zip",
                mime="application/zip",
                use_container_width=True
            )
