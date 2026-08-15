# excel_utils.py
import pandas as pd
import openpyxl
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.styles import Border, Side, Alignment, Font
import io
import re
from copy import copy

# ข้อมูลตารางเวรตั้งต้น
shift_data = {
    "ว": {"text": "(06.00-18.00) น.", "hours": 4}, "ค": {"text": "(00.00-06.00)(18.00-24.00) น.", "hours": 4},
    "ว/ค": {"text": "(06.00-12.00)(18.00-24.00) น.", "hours": 4}, "ค/ว": {"text": "(00.00-06.00)(12.00-18.00) น.", "hours": 4},
    "0-12": {"text": "(00.00-12.00) น.", "hours": 4}, "00-12": {"text": "(00.00-12.00) น.", "hours": 4},
    "12-24": {"text": "(12.00-24.00) น.", "hours": 4}, "00-24": {"text": "(00.00-24.00) น.", "hours": 4},
    "(ว)": {"text": "(06.00-18.00) น.", "hours": 4}, "(ค)": {"text": "(00.00-06.00)(18.00-24.00) น.", "hours": 4},
    "(ว/ค)": {"text": "(06.00-12.00)(18.00-24.00) น.", "hours": 4}, "(ค/ว)": {"text": "(00.00-06.00)(12.00-18.00) น.", "hours": 4},
    "(0-12)": {"text": "(00.00-12.00) น.", "hours": 4}, "(00-12)": {"text": "(00.00-12.00) น.", "hours": 4},
    "(12-24)": {"text": "(12.00-24.00) น.", "hours": 4},
    "ย": {"text": "ย.", "hours": "-"}, "ย.": {"text": "ย.", "hours": "-"},
    "พ": {"text": "พ.", "hours": "-"}, "พ.": {"text": "พ.", "hours": "-"},
    "ป": {"text": "ป.", "hours": "-"}, "ป.": {"text": "ป.", "hours": "-"},
    "ก": {"text": "ก.", "hours": "-"}, "ก.": {"text": "ก.", "hours": "-"},
    "น": {"text": "น.", "hours": "-"}, "น.": {"text": "น.", "hours": "-"},
    "ล": {"text": "ล.", "hours": "-"}, "ล.": {"text": "ล.", "hours": "-"}, "ลา": {"text": "ลา", "hours": "-"},
}
leave_types = ["ย", "ย.", "พ", "พ.", "ป", "ป.", "ก", "ก.", "น", "น.", "ล", "ล.", "ลา"]

def parse_days_count(day_str):
    total = 0
    if not day_str or day_str == "-": return 0
    parts = str(day_str).split(",")
    for p in parts:
        p = p.strip()
        if "-" in p:
            try:
                start, end = p.split("-")
                total += (int(end) - int(start)) + 1
            except: pass
        elif p.isdigit():
            total += 1
    return total

def parse_holiday_string_to_set(day_str):
    holiday_set = set()
    if not day_str or day_str == "-": return holiday_set
    parts = str(day_str).split(",")
    for p in parts:
        p = p.strip()
        if "-" in p:
            try:
                start, end = p.split("-")
                for d in range(int(start), int(end) + 1):
                    holiday_set.add(d)
            except: pass
        elif p.isdigit():
            holiday_set.add(int(p))
    return holiday_set

def extract_employee_stats(roster_row):
    weekly_days = []
    leave_days_vacation = []
    
    is_in_period = False
    for d in range(1, 32):
        val = str(roster_row.get(str(d), "")).strip()
        if "(" in val: is_in_period = True
        clean_val = val.replace("(", "").replace(")", "")
        
        if clean_val in ['ย', 'ย.']:
            weekly_days.append(d)
        elif is_in_period and clean_val and clean_val not in ["พ", "พ.", "ป", "ป.", "ก", "ก.", "น", "น.", "ล", "ล.", "ลา", "-"]:
            weekly_days.append(d)
            
        if clean_val in ['พ', 'พ.']:
            leave_days_vacation.append(d)
            
        if ")" in val: is_in_period = False

    def to_ranges(days_list):
        if not days_list: return []
        ranges = []
        start = days_list[0]
        prev = days_list[0]
        for d in days_list[1:]:
            if d == prev + 1: prev = d
            else:
                ranges.append(f"{start}-{prev}" if start != prev else f"{start}")
                start = d
                prev = d
        ranges.append(f"{start}-{prev}" if start != prev else f"{start}")
        return ranges
        
    w_ranges = to_ranges(weekly_days)
    val_4 = w_ranges[0] if len(w_ranges) > 0 else "-"
    val_5 = ",".join(w_ranges[1:]) if len(w_ranges) > 1 else "-"
    val_17 = f"{len(weekly_days):02d}"
    
    v_ranges = to_ranges(leave_days_vacation)
    val_9 = ",".join(v_ranges) if v_ranges else "-"
    val_10 = str(leave_days_vacation[0]) if leave_days_vacation else "-"
    val_11 = str(leave_days_vacation[-1]) if leave_days_vacation else "-"
    val_12 = f"{len(leave_days_vacation):02d}"
    
    return val_4, val_5, val_17, val_9, val_10, val_11, val_12

def generate_109(global_vars, roster_df, num_days, first_weekday):
    try: wb = openpyxl.load_workbook("109เปล่า.xlsx")
    except: return None, 0
    pages, current_page_rows = [], []
    for idx, row in roster_df.iterrows():
        if len(current_page_rows) >= 15 or (row.get("ขึ้นหน้าใหม่", False) and len(current_page_rows) > 0):
            pages.append(current_page_rows)
            current_page_rows = []
        current_page_rows.append(row)
    if current_page_rows: pages.append(current_page_rows)
    total_pages = max(len(pages), 1)
    if len(pages) == 0: pages = [[]]
    template_ws = wb.active
    template_ws.title = "หน้าที่ 1"
    worksheets = [template_ws]
    for p in range(2, total_pages + 1):
        new_ws = wb.copy_worksheet(template_ws)
        new_ws.title = f"หน้าที่ {p}"
        worksheets.append(new_ws)
    replacements_109 = {"[14]": global_vars["val_14"], "[13]": global_vars["val_13"], "[8]": global_vars["val_8"], "[7]": global_vars["val_7"]}
    days_th_abbr = ["จ.", "อ.", "พ.", "พฤ.", "ศ.", "ส.", "อา."]
    
    ph_dict_local = global_vars.get("public_holidays_dict", {})
    ph_append_str = ""
    if ph_dict_local:
        ph_texts = []
        for d_str, name in sorted(ph_dict_local.items(), key=lambda x: int(x[0])):
            ph_texts.append(f"วันที่ {d_str} {name}" if name else f"วันที่ {d_str}")
        ph_append_str = " (" + ", ".join(ph_texts) + ")"

    for page_idx, ws in enumerate(worksheets):
        page_num = page_idx + 1
        page_data = pages[page_idx]
        for r in range(1, 15):
            for c in range(1, 40):
                c_cell = ws.cell(row=r, column=c)
                val = c_cell.value
                if val and isinstance(val, str):
                    new_val = val
                    for k, v in replacements_109.items(): new_val = new_val.replace(k, str(v))
                    if "หน้า" in new_val and "/" in new_val: new_val = re.sub(r'หน้า\s*\d+\s*/\s*\d+', f'หน้า {page_num}/{total_pages}', new_val)
                    if type(c_cell).__name__ != 'MergedCell': c_cell.value = new_val
                    
        for r in range(35, 55):
            for c in range(1, 40):
                c_cell = ws.cell(row=r, column=c)
                val = c_cell.value
                if val and isinstance(val, str):
                    new_val = val
                    for k, v in replacements_109.items(): 
                        if k in new_val:
                            new_val = new_val.replace(k, str(v))
                    
                    if "วันหยุดนักขัตฤกษ์" in new_val and ph_append_str:
                        if ph_append_str not in new_val:
                            new_val = new_val.replace("วันหยุดนักขัตฤกษ์", f"วันหยุดนักขัตฤกษ์{ph_append_str}")
                            
                    if type(c_cell).__name__ != 'MergedCell': 
                        c_cell.value = new_val
        
        date_row = None
        for r in range(4, 10):
            if str(ws.cell(row=r, column=3).value).strip() == "1" and str(ws.cell(row=r, column=4).value).strip() == "2":
                date_row = r; break
        if not date_row: date_row = 7
        day_row = date_row - 1
        
        for r in range(8, 37, 2):
            ws.cell(row=r, column=1).value = ""
            ws.cell(row=r, column=2).value = ""
            ws.cell(row=r+1, column=2).value = ""
            for d in range(1, 32): ws.cell(row=r, column=2+d).value = ""
            
        for d in range(1, 32):
            col = 2 + d
            if d <= num_days:
                wd = (first_weekday + d - 1) % 7
                ws.cell(row=day_row, column=col).value = days_th_abbr[wd]
            else:
                ws.cell(row=day_row, column=col).value = ""
                ws.cell(row=date_row, column=col).value = ""
        
        current_excel_row = 8
        for row_data in page_data:
            ws.cell(row=current_excel_row, column=1).value = row_data['ลำดับ']
            ws.cell(row=current_excel_row, column=2).value = str(row_data['ชื่อ-สกุล']).strip()
            ws.cell(row=current_excel_row+1, column=2).value = str(row_data['ตำแหน่งเบิก']).strip()
            role_val = str(row_data.get('Role (หน้าที่)', '')).strip()
            
            for d in range(1, 32):
                col = 2 + d
                c_cell = ws.cell(row=current_excel_row, column=col)
                if d <= num_days:
                    shift = row_data.get(str(d), "")
                    c_cell.value = str(shift).strip() if pd.notna(shift) else ""
                    if c_cell.font: new_font = copy(c_cell.font)
                    else: new_font = Font()
                    
                    if shift:
                        s_clean = str(shift).strip().replace("(", "").replace(")", "")
                        if s_clean in leave_types: new_font.color = "FF0000" 
                        elif s_clean in ["00-12", "0-12", "12-24"]:
                            new_font.color = "008000" 
                            if new_font.size: new_font.size = new_font.size - 2 
                            else: new_font.size = 12
                        elif role_val == "กั้นถนนฯฉิมพลี" and s_clean in ["ว", "ค", "ว/ค", "ค/ว"]:
                            new_font.color = "0000FF" 
                        else: new_font.color = "000000" 
                    else: new_font.color = "000000"
                    c_cell.font = new_font
                else:
                    c_cell.value = ""
                    if c_cell.font:
                        nf = copy(c_cell.font)
                        nf.color = "000000"
                        c_cell.font = nf
                
            current_excel_row += 2
            
        if not ws.sheet_properties.pageSetUpPr: ws.sheet_properties.pageSetUpPr = PageSetupProperties()
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0 
        ws.page_setup.paperSize = ws.PAPERSIZE_A4; ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, total_pages

def generate_177(emp_info, roster_data, global_vars, ind_vars, num_days):
    if not emp_info: return None
    
    emp_id_str = str(emp_info.get("เลขประจำตัว", "-"))
    if emp_id_str.endswith(".0"): emp_id_str = emp_id_str[:-2]
    
    raw_salary = str(emp_info.get('เงินเดือน', '-'))
    if raw_salary != "-" and raw_salary.replace('.', '', 1).isdigit():
        salary_str = f"{float(raw_salary):,.0f}"
    else:
        salary_str = raw_salary

    try: wb = openpyxl.load_workbook("ใบ177 Update.xlsx")
    except: return None
    ws = wb.active
    replacements = {
        "[NAME]": emp_info["ชื่อ-สกุล"], "[16]": emp_info.get("รหัสบัญชี", "-"), "[15]": emp_info["ประเภทบัญชี"],
        "[14]": global_vars["val_14"], "[13]": global_vars["val_13"], 
        "[17]": ind_vars.get("val_17", ""), "[12]": ind_vars.get("val_12", ""),
        "[11]": ind_vars.get("val_11", ""), "[10]": ind_vars.get("val_10", ""), "[9]": ind_vars.get("val_9", ""),
        "[8]": global_vars["val_8"], "[7]": global_vars["val_7"], "[6]": ind_vars.get("val_6", ""),
        "[5]": ind_vars.get("val_5", ""), "[4]": ind_vars.get("val_4", ""),
        "[3]": salary_str,  
        "[2]": emp_id_str, "[1]": emp_info["ตำแหน่ง"]
    }
    for r in range(1, 55):
        for c in range(1, 40): 
            c_cell = ws.cell(row=r, column=c)
            val = c_cell.value
            if val and isinstance(val, str) and "[" in val:
                new_val = val
                for k, v in replacements.items(): new_val = new_val.replace(k, str(v))
                if type(c_cell).__name__ != 'MergedCell': c_cell.value = new_val
                
    start_row = 7
    rate_val = float(emp_info["เรท"]) if emp_info["เรท"] else 0.0
    rate_baht = int(rate_val)
    rate_satang = int(round((rate_val - rate_baht) * 100))
    
    def set_cell_val_color(r, c, val, color_hex="000000"):
        cell = ws.cell(row=r, column=c)
        if type(cell).__name__ != 'MergedCell':
            cell.value = val
            if cell.font:
                nf = copy(cell.font)
                nf.color = color_hex
                cell.font = nf
            else:
                cell.font = Font(color=color_hex)
    
    for day in range(1, 32):
        row = start_row + day - 1
        if day > num_days:
            set_cell_val_color(row, 2, "", "000000")
            for col in range(3, 8): set_cell_val_color(row, col, "", "000000")
            continue
            
        shift_raw = str(roster_data.get(str(day), "")).strip()
        shift_clean = shift_raw.replace("(", "").replace(")", "")
        sData = shift_data.get(shift_clean)
        
        is_holiday = shift_clean in leave_types
        font_color = "FF0000" if is_holiday else "000000"
        
        if sData and sData["hours"] != "-":
            hours_val = int(sData["hours"])
            total_money = hours_val * rate_val
            total_baht = int(total_money)
            total_satang = int(round((total_money - total_baht) * 100))

            set_cell_val_color(row, 2, sData["text"], font_color)
            set_cell_val_color(row, 3, hours_val, font_color)
            set_cell_val_color(row, 4, rate_baht if rate_baht > 0 else 0, font_color)
            set_cell_val_color(row, 5, f"{rate_satang:02d}", font_color)
            set_cell_val_color(row, 6, total_baht if total_baht > 0 else 0, font_color)
            set_cell_val_color(row, 7, f"{total_satang:02d}", font_color)
        else:
            val = sData["text"] if sData else (shift_raw if shift_raw else "-")
            set_cell_val_color(row, 2, val, font_color)
            for col in range(3, 8): 
                set_cell_val_color(row, col, "-", font_color)
                    
    for r in range(37, 45):
        cell_v1 = str(ws.cell(row=r, column=1).value).strip()
        cell_v2 = str(ws.cell(row=r, column=2).value).strip()
        if "รวม" in cell_v1 or "รวม" in cell_v2:
            set_cell_val_color(r, 4, rate_baht if rate_baht > 0 else 0, "000000")
            set_cell_val_color(r, 5, f"{rate_satang:02d}", "000000")
            break
            
    if not ws.sheet_properties.pageSetUpPr: ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1; ws.page_setup.fitToWidth = 1
    ws.page_setup.paperSize = ws.PAPERSIZE_A4; ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_178(emp_info, roster_data, global_vars, ind_vars, num_days):
    if not emp_info: return None
    
    emp_id_str = str(emp_info.get("เลขประจำตัว", "-"))
    if emp_id_str.endswith(".0"): emp_id_str = emp_id_str[:-2]
    
    raw_salary = str(emp_info.get('เงินเดือน', '-'))
    if raw_salary != "-" and raw_salary.replace('.', '', 1).isdigit():
        salary_str = f"{float(raw_salary):,.0f}"
    else:
        salary_str = raw_salary

    try: wb = openpyxl.load_workbook("178 อัพเดท.xlsx")
    except: return None
    ws = wb.active
    
    ph_dict_local = global_vars.get("public_holidays_dict", {})
    total_45 = parse_days_count(ind_vars.get('val_4', '0')) + parse_days_count(ind_vars.get('val_5', '0'))
    
    replacements = {
        "[NAME]": emp_info["ชื่อ-สกุล"], "[16]": emp_info.get("รหัสบัญชี2", "-"), "[15]": emp_info["ประเภทบัญชี"],
        "[14]": global_vars["val_14"], "[13]": global_vars["val_13"],
        "[8]": global_vars["val_8"], "[7]": global_vars["val_7"],
        "[17]": ind_vars.get("val_17", ""), "[12]": ind_vars.get("val_12", ""), 
        "[11]": ind_vars.get("val_11", ""), "[10]": ind_vars.get("val_10", ""), "[9]": ind_vars.get("val_9", ""),
        "[6]": ind_vars.get("val_6", ""), "[5]": ind_vars.get("val_5", ""), "[4]": ind_vars.get("val_4", ""),
        "[3]": salary_str,  
        "[2]": emp_id_str, "[1]": emp_info["ตำแหน่ง"]
    }
    
    for r in range(1, 55):
        for c in range(1, 40): 
            c_cell = ws.cell(row=r, column=c)
            val = c_cell.value
            if val and isinstance(val, str):
                new_val = val
                for k, v in replacements.items(): 
                    if k in new_val:
                        new_val = new_val.replace(k, str(v))
                if "รวม" in new_val and "วัน" in new_val:
                    new_val = re.sub(r'รวม\s*\d+\s*วัน', f'รวม {total_45:02d} วัน', new_val)
                if type(c_cell).__name__ != 'MergedCell': c_cell.value = new_val

    rate_val = float(emp_info["เรท"]) if emp_info["เรท"] else 0.0
    daily_rate = rate_val * 8
    
    if type(ws.cell(row=3, column=12)).__name__ != 'MergedCell': ws.cell(row=3, column=12).value = rate_val
    if type(ws.cell(row=5, column=12)).__name__ != 'MergedCell': ws.cell(row=5, column=12).value = daily_rate
        
    start_row = 7
    weekly_holiday_count = 0
    public_holiday_count = 0
    is_in_weekly_period = False
    
    manual_weekly_holidays = parse_holiday_string_to_set(ind_vars.get('val_4', '0')) | parse_holiday_string_to_set(ind_vars.get('val_5', '0'))
    
    for day in range(1, 32):
        row = start_row + day
        ws.cell(row=row, column=1).value = str(day) 
        
        for col in range(2, 11):
            if type(ws.cell(row=row, column=col)).__name__ != 'MergedCell':
                ws.cell(row=row, column=col).value = None

        if day > num_days:
            ws.cell(row=row, column=1).value = ""
            continue
            
        shift_raw = str(roster_data.get(str(day), "")).strip()
        if "(" in shift_raw: is_in_weekly_period = True
        shift_clean = shift_raw.replace("(", "").replace(")", "")
        
        if shift_clean and shift_clean not in leave_types and shift_clean != "-":
            is_public = str(day) in ph_dict_local
            is_weekly = day in manual_weekly_holidays
                
            if is_public or is_weekly:
                t1_start, t1_end, t2_start, t2_end = None, None, None, None
                if shift_clean == "ว": t1_start, t1_end = "06.00", "18.00"
                elif shift_clean == "ค": t1_start, t1_end, t2_start, t2_end = "00.00", "06.00", "18.00", "24.00"
                elif shift_clean == "ว/ค": t1_start, t1_end, t2_start, t2_end = "06.00", "12.00", "18.00", "24.00"
                elif shift_clean == "ค/ว": t1_start, t1_end, t2_start, t2_end = "00.00", "06.00", "12.00", "18.00"
                elif shift_clean in ["0-12", "00-12"]: t1_start, t1_end = "00.00", "12.00"
                elif shift_clean == "12-24": t1_start, t1_end = "12.00", "24.00"
                elif shift_clean == "00-24": t1_start, t1_end = "00.00", "24.00"
                else: t1_start, t1_end = shift_clean, ""
                
                ws.cell(row=row, column=3).value = emp_info["ตำแหน่ง"]
                if t1_start: ws.cell(row=row, column=4).value = t1_start
                if t1_end: ws.cell(row=row, column=5).value = t1_end
                if t2_start: ws.cell(row=row, column=6).value = t2_start
                if t2_end: ws.cell(row=row, column=7).value = t2_end
                ws.cell(row=row, column=8).value = 1 
                ws.cell(row=row, column=10).value = daily_rate 
                
                if is_public:
                    h_name = ph_dict_local.get(str(day), "วันหยุดนักขัตฤกษ์")
                    if not h_name.strip(): h_name = "วันหยุดนักขัตฤกษ์"
                    ws.cell(row=row, column=2).value = f"({h_name})"
                    public_holiday_count += 1
                elif is_weekly:
                    ws.cell(row=row, column=2).value = "(วันหยุดประจำสัปดาห์)"
                    weekly_holiday_count += 1
                    
        if ")" in shift_raw: is_in_weekly_period = False
                    
    if type(ws.cell(row=39, column=8)).__name__ != 'MergedCell': ws.cell(row=39, column=8).value = None
    if type(ws.cell(row=40, column=8)).__name__ != 'MergedCell': ws.cell(row=40, column=8).value = None
    
    total_days_final = weekly_holiday_count + public_holiday_count
    ws.cell(row=39, column=8).value = total_days_final
    
    if type(ws.cell(row=42, column=8)).__name__ != 'MergedCell':
        ws.cell(row=42, column=8).value = f"=SUM(H39:H41)"
    
    if not ws.sheet_properties.pageSetUpPr: ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1; ws.page_setup.fitToWidth = 1
    ws.page_setup.paperSize = ws.PAPERSIZE_A4; ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_report_work(emp_info, roster_data, global_vars, ind_vars, num_days):
    if not emp_info: return None
    
    emp_id_str = str(emp_info.get("เลขประจำตัว", "-"))
    if emp_id_str.endswith(".0"): emp_id_str = emp_id_str[:-2]
    
    raw_salary = str(emp_info.get('เงินเดือน', '-'))
    if raw_salary != "-" and raw_salary.replace('.', '', 1).isdigit():
        salary_str = f"{float(raw_salary):,.0f}"
    else:
        salary_str = raw_salary
        
    try: wb = openpyxl.load_workbook("รายงานปฏิบัติงาน.xlsx")
    except: return None
    ws = wb.active

    replacements = {
        "[NAME]": emp_info["ชื่อ-สกุล"],
        "[1]": emp_info["ตำแหน่ง"],
        "[2]": emp_id_str,
        "[3]": salary_str,
        "[14]": global_vars["val_14"],
        "[13]": global_vars["val_13"],
        "[8]": global_vars["val_8"], 
        "[7]": global_vars["val_7"],
        "[17]": ind_vars.get("val_17", ""), 
    }

    for r in range(1, 55):
        for c in range(1, 40):
            c_cell = ws.cell(row=r, column=c)
            val = c_cell.value
            if val and isinstance(val, str):
                new_val = val
                for k, v in replacements.items(): new_val = new_val.replace(k, str(v))
                if type(c_cell).__name__ != 'MergedCell': 
                    c_cell.value = new_val

    if type(ws.cell(row=2, column=7)).__name__ != 'MergedCell':
        ws.cell(row=2, column=7).value = global_vars["val_13"]
    if type(ws.cell(row=44, column=2)).__name__ != 'MergedCell':
        ws.cell(row=44, column=2).value = emp_info["ชื่อ-สกุล"]

    start_row = 8
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    ranges_to_unmerge = []
    for merged_range in list(ws.merged_cells.ranges):
        if start_row <= merged_range.min_row <= start_row + 31 and start_row <= merged_range.max_row <= start_row + 31:
            if merged_range.min_col >= 2 and merged_range.max_col <= 7:
                ranges_to_unmerge.append(merged_range.coord)
    
    for r_coord in ranges_to_unmerge:
        ws.unmerge_cells(r_coord)
        
    def apply_style(r, c, val, color_hex="000000"):
        cell = ws.cell(row=r, column=c)
        if type(cell).__name__ != 'MergedCell':
            cell.value = val
            cell.border = thin_border
            if cell.font:
                nf = copy(cell.font)
                nf.color = color_hex
                cell.font = nf
            else:
                cell.font = Font(color=color_hex)
            return cell
        return None

    for day in range(1, 32):
        row = start_row + day - 1
        
        if day > num_days:
            for c in range(2, 8): 
                cell = ws.cell(row=row, column=c)
                cell.value = None
                cell.border = thin_border
            apply_style(row, 8, "", "000000")
            continue
            
        shift_raw = str(roster_data.get(str(day), "")).strip()
        start_time, end_time = "-", "-"
        is_holiday = False
        
        if shift_raw:
            s_clean = shift_raw.replace("(", "").replace(")", "")
            if s_clean == "ว": start_time, end_time = "06.00", "18.00"
            elif s_clean == "ค": start_time, end_time = "00.00-06.00", "18.00-24.00"
            elif s_clean == "ว/ค": start_time, end_time = "06.00-12.00", "18.00-24.00"
            elif s_clean == "ค/ว": start_time, end_time = "00.00-06.00", "12.00-18.00"
            elif s_clean in ["0-12", "00-12"]: start_time, end_time = "00.00", "12.00"
            elif s_clean == "12-24": start_time, end_time = "12.00", "24.00"
            elif s_clean == "00-24": start_time, end_time = "00.00", "24.00"
            elif s_clean in leave_types: 
                start_time = s_clean
                end_time = ""
                is_holiday = True 
            else: start_time, end_time = shift_raw, "" 
            
        font_color = "FF0000" if is_holiday else "000000"

        for c in range(2, 8): 
            cell = ws.cell(row=row, column=c)
            cell.value = None
            cell.border = thin_border

        if is_holiday:
            cell = apply_style(row, 2, start_time, font_color)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
            if cell: cell.alignment = center_align
        else:
            cell1 = apply_style(row, 2, start_time, font_color)
            cell2 = apply_style(row, 5, end_time, font_color)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
            if cell1: cell1.alignment = center_align
            if cell2: cell2.alignment = center_align

        apply_style(row, 8, emp_info["ตำแหน่ง"] if start_time not in ["-", ""] and not is_holiday else "", "000000")
            
    if not ws.sheet_properties.pageSetUpPr: ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1; ws.page_setup.fitToWidth = 1
    ws.page_setup.paperSize = ws.PAPERSIZE_A4; ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
